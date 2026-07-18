"""
rich_extractor.py
=================
Extracts "rich content" from spreadsheet/document files that the plain
DataFrame pipeline would discard:

  • Excel charts  → rendered as PNG bytes via matplotlib
  • Excel formulas → formula strings per cell (row, col)
  • LaTeX / math notation in cells → rendered as PNG images
  • OMML equations in DOCX/PPTX → rendered as PNG images
  • Chemical/scientific notation → preserved via proper Unicode handling

All methods are self-contained and return serialisable types (bytes, dicts,
lists of bytes) so they can be passed across thread boundaries.
"""

from __future__ import annotations
import io
import logging
import re
import unicodedata
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# ── LaTeX / math patterns ─────────────────────────────────────────────────────
_LATEX_INLINE  = re.compile(r'\$(?!\$).+?\$', re.DOTALL)


def _has_math(text: str) -> bool:
    """Quick check: does this string look like it contains math/LaTeX?"""
    if not text:
        return False
    return bool(
        _LATEX_DISPLAY.search(text) or
        _LATEX_INLINE.search(text) or
        _LATEX_CMDS.search(text)
    )


def _strip_latex_delimiters(expr: str) -> str:
    """Remove surrounding $ or $$ delimiters."""
    expr = expr.strip()
    if expr.startswith("$$") and expr.endswith("$$"):
        return expr[2:-2].strip()
    if expr.startswith("$") and expr.endswith("$"):
        return expr[1:-1].strip()
    return expr


# ── Matplotlib helpers ────────────────────────────────────────────────────────

def _mpl_backend() -> None:
    """Ensure a non-interactive Agg backend is active."""
    import matplotlib
    matplotlib.use("Agg")


def render_math_png(expr: str, fontsize: int = 14, dpi: int = 150) -> bytes | None:
    """
    Render a mathematical expression (LaTeX) to PNG bytes.
    Returns None if rendering fails.
    """
    try:
        _mpl_backend()
        import matplotlib.pyplot as plt

        clean = _strip_latex_delimiters(expr)
        # Wrap in $…$ if not already a display environment
        if not (clean.startswith(r"\begin") or clean.startswith(r"\[")):
            latex = f"${clean}$"
        else:
            latex = clean

        fig = plt.figure(figsize=(6, 1.2))
        fig.patch.set_facecolor("white")
        t = fig.text(
            0.5, 0.5, latex,
            ha="center", va="center",
            fontsize=fontsize,
            color="black",
        )
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=dpi,
                    bbox_inches="tight", pad_inches=0.15,
                    facecolor="white")
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        logger.debug(f"Math render failed for '{expr[:40]}': {e}")
        return None


# ── Excel chart extraction ────────────────────────────────────────────────────

def _resolve_ref(ws, ref) -> list:
    """Resolve an openpyxl Reference object to a flat list of values."""
    if ref is None:
        return []
    vals: list = []
    try:
        for row in ws.iter_rows(
            min_row=ref.min_row, max_row=ref.max_row,
            min_col=ref.min_col, max_col=ref.max_col,
            values_only=True,
        ):
            for v in row:
                if v is not None:
                    vals.append(v)
    except Exception:
        pass
    return vals


def _chart_title(chart) -> str:
    """Best-effort chart title extraction."""
    try:
        t = chart.title
        if t is None:
            return ""
        if isinstance(t, str):
            return t
        # openpyxl Title object
        if hasattr(t, "tx") and t.tx:
            tx = t.tx
            if hasattr(tx, "rich") and tx.rich:
                parts = []
                for para in tx.rich.paragraphs:
                    for run in para.runs:
                        if run.t:
                            parts.append(run.t)
                return "".join(parts)
            if hasattr(tx, "strRef") and tx.strRef and tx.strRef.f:
                return tx.strRef.f
        return ""
    except Exception:
        return ""


def _series_values(ws, series) -> tuple[list, list, str]:
    """
    Return (categories, values, series_label) from an openpyxl series object.
    Works for BarChart, LineChart, AreaChart, PieChart series.
    """
    cats: list = []
    vals: list = []
    label = ""
    try:
        # Series title / legend entry
        if hasattr(series, "title") and series.title:
            st = series.title
            if hasattr(st, "v") and st.v:
                label = str(st.v)
            elif hasattr(st, "strRef") and st.strRef:
                f = getattr(st.strRef, "f", "")
                label = f.split("!")[-1].strip('"') if "!" in f else f
        # Values
        val_ref = getattr(series, "val", None) or getattr(series, "yVal", None)
        if val_ref and hasattr(val_ref, "numRef") and val_ref.numRef:
            nr = val_ref.numRef
            if hasattr(nr, "ref") and nr.ref:
                from openpyxl.utils import range_to_tuple
                try:
                    ws2 = ws.parent[nr.ref.split("!")[0].strip("'")] \
                        if "!" in nr.ref else ws
                    ref_str = nr.ref.split("!")[-1] if "!" in nr.ref else nr.ref
                    from openpyxl import load_workbook as _lw
                    # Use cached values if available
                    if nr.numCache and nr.numCache.pt:
                        vals = [float(p.v) for p in nr.numCache.pt if p.v is not None]
                    else:
                        from openpyxl.utils.cell import range_boundaries
                        mn_col, mn_row, mx_col, mx_row = range_boundaries(ref_str)
                        for row in ws.iter_rows(
                            min_row=mn_row, max_row=mx_row,
                            min_col=mn_col, max_col=mx_col,
                            values_only=True,
                        ):
                            for v in row:
                                if isinstance(v, (int, float)):
                                    vals.append(float(v))
                except Exception:
                    pass
        # Categories
        cat_ref = getattr(series, "cat", None) or getattr(series, "xVal", None)
        if cat_ref:
            if hasattr(cat_ref, "numRef") and cat_ref.numRef and cat_ref.numRef.numCache:
                cats = [str(p.v) for p in cat_ref.numRef.numCache.pt if p.v is not None]
            elif hasattr(cat_ref, "strRef") and cat_ref.strRef and cat_ref.strRef.strCache:
                cats = [str(p.v) for p in cat_ref.strRef.strCache.pt if p.v is not None]
    except Exception as e:
        logger.debug(f"Series extraction error: {e}")
    return cats, vals, label


def _render_chart(chart, ws) -> bytes | None:
    """Render a single openpyxl chart object to PNG bytes using matplotlib."""
    try:
        _mpl_backend()
        import matplotlib.pyplot as plt
        import matplotlib.cm as cm
        import numpy as np

        chart_type = type(chart).__name__.lower()  # barchart, linechart, piechart …
        title = _chart_title(chart)
        series_list = list(chart.series)

        if not series_list:
            return None

        fig, ax = plt.subplots(figsize=(7, 4))
        fig.patch.set_facecolor("white")
        ax.set_facecolor("#f8f9fa")
        if title:
            ax.set_title(title, fontsize=11, fontweight="bold", pad=8)

        colors = cm.tab10.colors

        if "pie" in chart_type:
            cats, vals, _ = _series_values(ws, series_list[0])
            if not vals:
                plt.close(fig)
                return None
            labels = cats if cats else [f"Item {i+1}" for i in range(len(vals))]
            ax.pie(vals, labels=labels, autopct="%1.1f%%",
                   colors=colors[:len(vals)], startangle=90)

        elif "scatter" in chart_type or "bubble" in chart_type:
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                x = list(range(len(vals))) if not cats else cats
                try:
                    x = [float(v) for v in x]
                except (ValueError, TypeError):
                    x = list(range(len(vals)))
                if vals:
                    ax.scatter(x, vals, label=label or f"Series {idx+1}",
                               color=colors[idx % len(colors)], s=60, alpha=0.8)
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        elif "line" in chart_type or "area" in chart_type or "stock" in chart_type:
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                x = list(range(len(vals))) if not cats else cats
                if vals:
                    if "area" in chart_type:
                        ax.fill_between(range(len(vals)), vals,
                                        alpha=0.4, color=colors[idx % len(colors)])
                    ax.plot(vals, label=label or f"Series {idx+1}",
                            color=colors[idx % len(colors)], linewidth=2, marker="o",
                            markersize=4)
                    if cats:
                        step = max(1, len(cats) // 10)
                        ax.set_xticks(range(0, len(cats), step))
                        ax.set_xticklabels(
                            [str(cats[i]) for i in range(0, len(cats), step)],
                            rotation=30, fontsize=7, ha="right"
                        )
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        elif "radar" in chart_type:
            cats, vals, _ = _series_values(ws, series_list[0])
            if vals and cats:
                angles = np.linspace(0, 2 * np.pi, len(vals), endpoint=False).tolist()
                angles += angles[:1]
                vals_closed = vals + vals[:1]
                ax_r = fig.add_subplot(111, polar=True)
                ax.remove()
                ax_r.plot(angles, vals_closed, linewidth=1.5, color=colors[0])
                ax_r.fill(angles, vals_closed, alpha=0.25, color=colors[0])
                ax_r.set_xticks(angles[:-1])
                ax_r.set_xticklabels(cats, fontsize=8)

        else:
            # Default: grouped bar chart
            all_cats: list[str] = []
            all_vals: list[tuple[list, str]] = []
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                if vals:
                    all_vals.append((vals, label or f"Series {idx+1}"))
                    if cats and not all_cats:
                        all_cats = [str(c) for c in cats]

            if not all_vals:
                plt.close(fig)
                return None

            n_groups = max(len(v) for v, _ in all_vals)
            n_series = len(all_vals)
            x = np.arange(n_groups)
            width = 0.8 / n_series

            for idx, (vals, label) in enumerate(all_vals):
                offset = (idx - n_series / 2 + 0.5) * width
                bars = ax.bar(x[:len(vals)] + offset, vals, width,
                              label=label, color=colors[idx % len(colors)], alpha=0.85)

            if all_cats:
                step = max(1, len(all_cats) // 12)
                ax.set_xticks(x[::step])
                ax.set_xticklabels(
                    [all_cats[i] for i in range(0, len(all_cats), step)],
                    rotation=30, fontsize=7, ha="right",
                )
            if n_series > 1:
                ax.legend(fontsize=8)
            ax.grid(axis="y", alpha=0.3)

        plt.tight_layout(pad=0.8)
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=130, bbox_inches="tight",
                    facecolor="white")
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        logger.warning(f"Chart render failed [{type(chart).__name__}]: {e}")
        try:
            plt.close("all")
        except Exception:
            pass
        return None


# ── Public API ────────────────────────────────────────────────────────────────

def extract_excel_charts(path: str, sheet_name: str | None = None) -> list[dict]:
    """
    Extract all charts from an Excel worksheet.

    Returns a list of dicts:
        {
          "title": str,
          "png":   bytes,      # rendered PNG image
        }
    Returns an empty list on any failure.
    """
    results: list[dict] = []
    try:
        import openpyxl
        # Load with data_only=True so cached values are available for reference resolution
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet_names = wb.sheetnames if sheet_name is None else [sheet_name]
        for sn in sheet_names:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            charts = getattr(ws, "_charts", [])
            for chart in charts:
                png = _render_chart(chart, ws)
                if png:
                    results.append({
                        "title": _chart_title(chart) or sn,
                        "png": png,
                    })
        wb.close()
    except Exception as e:
        logger.warning(f"Excel chart extraction failed: {e}")
    return results


def extract_excel_formulas(path: str, sheet_name: str | None = None) -> dict[str, dict]:
    """
    Read formula strings from Excel cells (loads WITHOUT data_only).

    Returns:
        { sheet_name: { "A1": "=SUM(B1:B10)", ... } }
    """
    result: dict[str, dict] = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=False)
        names = wb.sheetnames if sheet_name is None else [sheet_name]
        for sn in names:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            sheet_formulas: dict[str, str] = {}
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        sheet_formulas[cell.coordinate] = v
            if sheet_formulas:
                result[sn] = sheet_formulas
        wb.close()
    except Exception as e:
        logger.debug(f"Formula extraction failed: {e}")
    return result


def detect_math_cells(data: dict[str, Any]) -> dict[str, dict[str, bytes]]:
    """
    Scan a dict of {sheet_name: DataFrame} and detect cells with LaTeX/math content.

    Returns:
        { sheet_name: { "R<row>C<col>": <png_bytes> } }
    """
    results: dict[str, dict[str, bytes]] = {}
    for sheet_name, df in data.items():
        sheet_imgs: dict[str, bytes] = {}
        for r_idx, row in enumerate(df.iter_rows(named=False)):
            for c_idx, val in enumerate(row):
                text = str(val) if val is not None else ""
                if _has_math(text):
                    png = render_math_png(text)
                    if png:
                        sheet_imgs[f"R{r_idx}C{c_idx}"] = png
        if sheet_imgs:
            results[sheet_name] = sheet_imgs
    return results


def extract_docx_equations(path: str) -> list[dict]:
    """
    Extract OMML (Office Math Markup Language) equations from a DOCX file.
    Returns list of { "text": str, "png": bytes | None }.
    """
    results: list[dict] = []
    try:
        from lxml import etree
        from zipfile import ZipFile

        NS_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
        NS_M = "http://schemas.openxmlformats.org/officeDocument/2006/math"

        with ZipFile(path) as zf:
            if "word/document.xml" not in zf.namelist():
                return results
            xml_bytes = zf.read("word/document.xml")

        root = etree.fromstring(xml_bytes)

        for omath in root.iter(f"{{{NS_M}}}oMath"):
            # Best-effort text extraction from OMML
            text_parts = []
            for t_elem in omath.iter(f"{{{NS_M}}}t"):
                if t_elem.text:
                    text_parts.append(t_elem.text)
            # Also grab any plain <w:t> inside runs inside the math
            for t_elem in omath.iter(f"{{{NS_W}}}t"):
                if t_elem.text:
                    text_parts.append(t_elem.text)

            raw_text = " ".join(text_parts).strip()
            if not raw_text:
                continue

            # Try to render as LaTeX/math expression
            png = render_math_png(raw_text) if raw_text else None
            results.append({"text": raw_text, "png": png})

    except Exception as e:
        logger.debug(f"DOCX equation extraction failed: {e}")
    return results


def extract_pptx_equations(path: str) -> list[dict]:
    """
    Extract math equations from a PPTX file.
    Returns list of { "slide": int, "text": str, "png": bytes | None }.
    """
    results: list[dict] = []
    try:
        from lxml import etree
        from zipfile import ZipFile

        NS_A14 = "http://schemas.microsoft.com/office/drawing/2010/main"
        NS_M   = "http://schemas.openxmlformats.org/officeDocument/2006/math"

        with ZipFile(path) as zf:
            slide_files = sorted(
                n for n in zf.namelist()
                if n.startswith("ppt/slides/slide") and n.endswith(".xml")
            )
            for slide_idx, slide_file in enumerate(slide_files, 1):
                xml_bytes = zf.read(slide_file)
                root = etree.fromstring(xml_bytes)
                for omath in root.iter(f"{{{NS_M}}}oMath"):
                    text_parts = [
                        t.text for t in omath.iter(f"{{{NS_M}}}t")
                        if t.text
                    ]
                    raw_text = " ".join(text_parts).strip()
                    if raw_text:
                        png = render_math_png(raw_text)
                        results.append({
                            "slide": slide_idx,
                            "text": raw_text,
                            "png": png,
                        })
    except Exception as e:
        logger.debug(f"PPTX equation extraction failed: {e}")
    return results


def enrich_pdf_with_charts(
    doc,  # fitz.Document (already built, will get new pages appended)
    charts: list[dict],
    title_prefix: str = "المخططات البيانية",
    equations: list[dict] | None = None,
) -> None:
    """
    Append a new page (or pages) to a PyMuPDF document containing
    chart images and optional equation images.
    """
    if not charts and not equations:
        return

    try:
        import pymupdf as fitz
        FONT = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
        PAGE_W, PAGE_H = 595, 842  # A4 portrait
        MARGIN = 30
        TITLE_H = 24
        GAP = 12

        page = doc.new_page(width=PAGE_W, height=PAGE_H)
        try:
            page.insert_font(fontname="dvb", fontfile=FONT)
        except Exception:
            pass

        y = MARGIN

        # Section title
        if charts:
            page.draw_rect(
                fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + TITLE_H),
                color=None, fill=(79/255, 70/255, 229/255), overlay=True,
            )
            page.insert_textbox(
                fitz.Rect(MARGIN + 6, y + 4, PAGE_W - MARGIN - 6, y + TITLE_H - 4),
                title_prefix,
                fontname="dvb", fontsize=11, color=(1, 1, 1),
                align=1,  # center
                overlay=True,
            )
            y += TITLE_H + GAP

        for ch in charts:
            png = ch.get("png")
            if not png:
                continue
            ch_title = ch.get("title", "")

            # If we're near the bottom, start a new page
            if y + 180 > PAGE_H - MARGIN:
                page = doc.new_page(width=PAGE_W, height=PAGE_H)
                try:
                    page.insert_font(fontname="dvb", fontfile=FONT)
                except Exception:
                    pass
                y = MARGIN

            # Chart title label
            if ch_title:
                page.insert_textbox(
                    fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + 14),
                    ch_title,
                    fontname="dvb", fontsize=9, color=(0.2, 0.2, 0.2),
                    align=1, overlay=True,
                )
                y += 16

            # Embed chart image
            available_w = PAGE_W - 2 * MARGIN
            img_stream = fitz.open(stream=png, filetype="png")
            img_w = img_stream[0].rect.width
            img_h = img_stream[0].rect.height
            img_stream.close()

            scale = min(available_w / max(img_w, 1), 200 / max(img_h, 1))
            draw_w = img_w * scale
            draw_h = img_h * scale
            x0 = MARGIN + (available_w - draw_w) / 2

            rect = fitz.Rect(x0, y, x0 + draw_w, y + draw_h)
            page.insert_image(rect, stream=png, overlay=True)
            y += draw_h + GAP

        # Equations section
        if equations:
            if y + TITLE_H + 20 > PAGE_H - MARGIN:
                page = doc.new_page(width=PAGE_W, height=PAGE_H)
                try:
                    page.insert_font(fontname="dvb", fontfile=FONT)
                except Exception:
                    pass
                y = MARGIN

            page.draw_rect(
                fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + TITLE_H),
                color=None, fill=(30/255, 100/255, 60/255), overlay=True,
            )
            page.insert_textbox(
                fitz.Rect(MARGIN + 6, y + 4, PAGE_W - MARGIN - 6, y + TITLE_H - 4),
                "المعادلات والصيغ الرياضية",
                fontname="dvb", fontsize=11, color=(1, 1, 1),
                align=1, overlay=True,
            )
            y += TITLE_H + GAP

            for eq in equations:
                png = eq.get("png")
                text = eq.get("text", "")
                if not png and not text:
                    continue

                if y + 50 > PAGE_H - MARGIN:
                    page = doc.new_page(width=PAGE_W, height=PAGE_H)
                    try:
                        page.insert_font(fontname="dvb", fontfile=FONT)
                    except Exception:
                        pass
                    y = MARGIN

                if png:
                    img_stream = fitz.open(stream=png, filetype="png")
                    img_w = img_stream[0].rect.width
                    img_h = img_stream[0].rect.height
                    img_stream.close()
                    available_w = PAGE_W - 2 * MARGIN
                    scale = min(available_w / max(img_w, 1), 60 / max(img_h, 1))
                    draw_w = img_w * scale
                    draw_h = img_h * scale
                    x0 = MARGIN + (available_w - draw_w) / 2
                    rect = fitz.Rect(x0, y, x0 + draw_w, y + draw_h)
                    page.insert_image(rect, stream=png, overlay=True)
                    y += draw_h + 4
                else:
                    page.insert_textbox(
                        fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + 18),
                        text,
                        fontname="dvb", fontsize=8, color=(0.1, 0.1, 0.5),
                        align=0, overlay=True,
                    )
                    y += 20

    except Exception as e:
        logger.warning(f"enrich_pdf_with_charts failed: {e}")


def charts_to_html_blocks(charts: list[dict], equations: list[dict] | None = None) -> str:
    """
    Convert extracted charts and equations to HTML <img> tags (base64 PNG).
    Returns an HTML string fragment ready to inject into an HTML page.
    """
    import base64
    parts: list[str] = []

    if charts:
        parts.append(
            '<div style="margin:24px 0;"><h3 style="font-family:sans-serif;'
            'color:#4F46E5;border-bottom:2px solid #4F46E5;padding-bottom:6px;">'
            '📊 المخططات البيانية</h3>'
        )
        for ch in charts:
            png = ch.get("png")
            title = ch.get("title", "")
            if not png:
                continue
            b64 = base64.b64encode(png).decode()
            parts.append(
                f'<figure style="margin:16px 0;text-align:center;">'
                f'<img src="data:image/png;base64,{b64}" '
                f'style="max-width:100%;border:1px solid #e2e8f0;border-radius:8px;'
                f'box-shadow:0 2px 8px rgba(0,0,0,.08);" alt="{title}"/>'
                f'{"<figcaption style=\"font-family:sans-serif;font-size:13px;color:#64748b;margin-top:6px;\">" + title + "</figcaption>" if title else ""}'
                f'</figure>'
            )
        parts.append('</div>')

    if equations:
        parts.append(
            '<div style="margin:24px 0;"><h3 style="font-family:sans-serif;'
            'color:#059669;border-bottom:2px solid #059669;padding-bottom:6px;">'
            '∑ المعادلات والصيغ الرياضية</h3>'
        )
        for eq in equations:
            png = eq.get("png")
            text = eq.get("text", "")
            if png:
                b64 = base64.b64encode(png).decode()
                parts.append(
                    f'<figure style="margin:12px 0;text-align:center;">'
                    f'<img src="data:image/png;base64,{b64}" '
                    f'style="max-height:60px;" alt="{text}"/>'
                    f'</figure>'
                )
            elif text:
                parts.append(
                    f'<div style="font-family:monospace;font-size:13px;'
                    f'background:#f1f5f9;padding:8px 12px;border-radius:6px;'
                    f'margin:8px 0;">{text}</div>'
                )
        parts.append('</div>')

    return "\n".join(parts)
_LATEX_DISPLAY = re.compile(r'\$\$.+?\$\$', re.DOTALL)


def _has_math(text: str) -> bool:
    """Quick check: does this string look like it contains math/LaTeX?"""
    if not text:
        return False
    return bool(
        _LATEX_DISPLAY.search(text) or
        _LATEX_INLINE.search(text) or
        _LATEX_CMDS.search(text)
    )


def _strip_latex_delimiters(expr: str) -> str:
    """Remove surrounding $ or $$ delimiters."""
    expr = expr.strip()
    if expr.startswith("$$") and expr.endswith("$$"):
        return expr[2:-2].strip()
    if expr.startswith("$") and expr.endswith("$"):
        return expr[1:-1].strip()
    return expr


# ── Matplotlib helpers ────────────────────────────────────────────────────────

def _mpl_backend() -> None:
    """Ensure a non-interactive Agg backend is active."""
    import matplotlib
    matplotlib.use("Agg")


def render_math_png(expr: str, fontsize: int = 14, dpi: int = 150) -> bytes | None:
    """
    Render a mathematical expression (LaTeX) to PNG bytes.
    Returns None if rendering fails.
    """
    try:
        _mpl_backend()
        import matplotlib.pyplot as plt

        clean = _strip_latex_delimiters(expr)
        # Wrap in $…$ if not already a display environment
        if not (clean.startswith(r"\begin") or clean.startswith(r"\[")):
            latex = f"${clean}$"
        else:
            latex = clean

        fig = plt.figure(figsize=(6, 1.2))
        fig.patch.set_facecolor("white")
        t = fig.text(
            0.5, 0.5, latex,
            ha="center", va="center",
            fontsize=fontsize,
            color="black",
        )
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=dpi,
                    bbox_inches="tight", pad_inches=0.15,
                    facecolor="white")
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        logger.debug(f"Math render failed for '{expr[:40]}': {e}")
        return None


# ── Excel chart extraction ────────────────────────────────────────────────────

def _resolve_ref(ws, ref) -> list:
    """Resolve an openpyxl Reference object to a flat list of values."""
    if ref is None:
        return []
    vals: list = []
    try:
        for row in ws.iter_rows(
            min_row=ref.min_row, max_row=ref.max_row,
            min_col=ref.min_col, max_col=ref.max_col,
            values_only=True,
        ):
            for v in row:
                if v is not None:
                    vals.append(v)
    except Exception:
        pass
    return vals


def _chart_title(chart) -> str:
    """Best-effort chart title extraction."""
    try:
        t = chart.title
        if t is None:
            return ""
        if isinstance(t, str):
            return t
        # openpyxl Title object
        if hasattr(t, "tx") and t.tx:
            tx = t.tx
            if hasattr(tx, "rich") and tx.rich:
                parts = []
                for para in tx.rich.paragraphs:
                    for run in para.runs:
                        if run.t:
                            parts.append(run.t)
                return "".join(parts)
            if hasattr(tx, "strRef") and tx.strRef and tx.strRef.f:
                return tx.strRef.f
        return ""
    except Exception:
        return ""


def _series_values(ws, series) -> tuple[list, list, str]:
    """
    Return (categories, values, series_label) from an openpyxl series object.
    Works for BarChart, LineChart, AreaChart, PieChart series.
    """
    cats: list = []
    vals: list = []
    label = ""
    try:
        # Series title / legend entry
        if hasattr(series, "title") and series.title:
            st = series.title
            if hasattr(st, "v") and st.v:
                label = str(st.v)
            elif hasattr(st, "strRef") and st.strRef:
                f = getattr(st.strRef, "f", "")
                label = f.split("!")[-1].strip('"') if "!" in f else f
        # Values
        val_ref = getattr(series, "val", None) or getattr(series, "yVal", None)
        if val_ref and hasattr(val_ref, "numRef") and val_ref.numRef:
            nr = val_ref.numRef
            if hasattr(nr, "ref") and nr.ref:
                from openpyxl.utils import range_to_tuple
                try:
                    ws2 = ws.parent[nr.ref.split("!")[0].strip("'")] \
                        if "!" in nr.ref else ws
                    ref_str = nr.ref.split("!")[-1] if "!" in nr.ref else nr.ref
                    from openpyxl import load_workbook as _lw
                    # Use cached values if available
                    if nr.numCache and nr.numCache.pt:
                        vals = [float(p.v) for p in nr.numCache.pt if p.v is not None]
                    else:
                        from openpyxl.utils.cell import range_boundaries
                        mn_col, mn_row, mx_col, mx_row = range_boundaries(ref_str)
                        for row in ws.iter_rows(
                            min_row=mn_row, max_row=mx_row,
                            min_col=mn_col, max_col=mx_col,
                            values_only=True,
                        ):
                            for v in row:
                                if isinstance(v, (int, float)):
                                    vals.append(float(v))
                except Exception:
                    pass
        # Categories
        cat_ref = getattr(series, "cat", None) or getattr(series, "xVal", None)
        if cat_ref:
            if hasattr(cat_ref, "numRef") and cat_ref.numRef and cat_ref.numRef.numCache:
                cats = [str(p.v) for p in cat_ref.numRef.numCache.pt if p.v is not None]
            elif hasattr(cat_ref, "strRef") and cat_ref.strRef and cat_ref.strRef.strCache:
                cats = [str(p.v) for p in cat_ref.strRef.strCache.pt if p.v is not None]
    except Exception as e:
        logger.debug(f"Series extraction error: {e}")
    return cats, vals, label


def _render_chart(chart, ws) -> bytes | None:
    """Render a single openpyxl chart object to PNG bytes using matplotlib."""
    try:
        _mpl_backend()
        import matplotlib.pyplot as plt
        import matplotlib.cm as cm
        import numpy as np

        chart_type = type(chart).__name__.lower()  # barchart, linechart, piechart …
        title = _chart_title(chart)
        series_list = list(chart.series)

        if not series_list:
            return None

        fig, ax = plt.subplots(figsize=(7, 4))
        fig.patch.set_facecolor("white")
        ax.set_facecolor("#f8f9fa")
        if title:
            ax.set_title(title, fontsize=11, fontweight="bold", pad=8)

        colors = cm.tab10.colors

        if "pie" in chart_type:
            cats, vals, _ = _series_values(ws, series_list[0])
            if not vals:
                plt.close(fig)
                return None
            labels = cats if cats else [f"Item {i+1}" for i in range(len(vals))]
            ax.pie(vals, labels=labels, autopct="%1.1f%%",
                   colors=colors[:len(vals)], startangle=90)

        elif "scatter" in chart_type or "bubble" in chart_type:
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                x = list(range(len(vals))) if not cats else cats
                try:
                    x = [float(v) for v in x]
                except (ValueError, TypeError):
                    x = list(range(len(vals)))
                if vals:
                    ax.scatter(x, vals, label=label or f"Series {idx+1}",
                               color=colors[idx % len(colors)], s=60, alpha=0.8)
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        elif "line" in chart_type or "area" in chart_type or "stock" in chart_type:
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                x = list(range(len(vals))) if not cats else cats
                if vals:
                    if "area" in chart_type:
                        ax.fill_between(range(len(vals)), vals,
                                        alpha=0.4, color=colors[idx % len(colors)])
                    ax.plot(vals, label=label or f"Series {idx+1}",
                            color=colors[idx % len(colors)], linewidth=2, marker="o",
                            markersize=4)
                    if cats:
                        step = max(1, len(cats) // 10)
                        ax.set_xticks(range(0, len(cats), step))
                        ax.set_xticklabels(
                            [str(cats[i]) for i in range(0, len(cats), step)],
                            rotation=30, fontsize=7, ha="right"
                        )
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        elif "radar" in chart_type:
            cats, vals, _ = _series_values(ws, series_list[0])
            if vals and cats:
                angles = np.linspace(0, 2 * np.pi, len(vals), endpoint=False).tolist()
                angles += angles[:1]
                vals_closed = vals + vals[:1]
                ax_r = fig.add_subplot(111, polar=True)
                ax.remove()
                ax_r.plot(angles, vals_closed, linewidth=1.5, color=colors[0])
                ax_r.fill(angles, vals_closed, alpha=0.25, color=colors[0])
                ax_r.set_xticks(angles[:-1])
                ax_r.set_xticklabels(cats, fontsize=8)

        else:
            # Default: grouped bar chart
            all_cats: list[str] = []
            all_vals: list[tuple[list, str]] = []
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                if vals:
                    all_vals.append((vals, label or f"Series {idx+1}"))
                    if cats and not all_cats:
                        all_cats = [str(c) for c in cats]

            if not all_vals:
                plt.close(fig)
                return None

            n_groups = max(len(v) for v, _ in all_vals)
            n_series = len(all_vals)
            x = np.arange(n_groups)
            width = 0.8 / n_series

            for idx, (vals, label) in enumerate(all_vals):
                offset = (idx - n_series / 2 + 0.5) * width
                bars = ax.bar(x[:len(vals)] + offset, vals, width,
                              label=label, color=colors[idx % len(colors)], alpha=0.85)

            if all_cats:
                step = max(1, len(all_cats) // 12)
                ax.set_xticks(x[::step])
                ax.set_xticklabels(
                    [all_cats[i] for i in range(0, len(all_cats), step)],
                    rotation=30, fontsize=7, ha="right",
                )
            if n_series > 1:
                ax.legend(fontsize=8)
            ax.grid(axis="y", alpha=0.3)

        plt.tight_layout(pad=0.8)
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=130, bbox_inches="tight",
                    facecolor="white")
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        logger.warning(f"Chart render failed [{type(chart).__name__}]: {e}")
        try:
            plt.close("all")
        except Exception:
            pass
        return None


# ── Public API ────────────────────────────────────────────────────────────────

def extract_excel_charts(path: str, sheet_name: str | None = None) -> list[dict]:
    """
    Extract all charts from an Excel worksheet.

    Returns a list of dicts:
        {
          "title": str,
          "png":   bytes,      # rendered PNG image
        }
    Returns an empty list on any failure.
    """
    results: list[dict] = []
    try:
        import openpyxl
        # Load with data_only=True so cached values are available for reference resolution
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet_names = wb.sheetnames if sheet_name is None else [sheet_name]
        for sn in sheet_names:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            charts = getattr(ws, "_charts", [])
            for chart in charts:
                png = _render_chart(chart, ws)
                if png:
                    results.append({
                        "title": _chart_title(chart) or sn,
                        "png": png,
                    })
        wb.close()
    except Exception as e:
        logger.warning(f"Excel chart extraction failed: {e}")
    return results


def extract_excel_formulas(path: str, sheet_name: str | None = None) -> dict[str, dict]:
    """
    Read formula strings from Excel cells (loads WITHOUT data_only).

    Returns:
        { sheet_name: { "A1": "=SUM(B1:B10)", ... } }
    """
    result: dict[str, dict] = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=False)
        names = wb.sheetnames if sheet_name is None else [sheet_name]
        for sn in names:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            sheet_formulas: dict[str, str] = {}
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        sheet_formulas[cell.coordinate] = v
            if sheet_formulas:
                result[sn] = sheet_formulas
        wb.close()
    except Exception as e:
        logger.debug(f"Formula extraction failed: {e}")
    return result


def detect_math_cells(data: dict[str, Any]) -> dict[str, dict[str, bytes]]:
    """
    Scan a dict of {sheet_name: DataFrame} and detect cells with LaTeX/math content.

    Returns:
        { sheet_name: { "R<row>C<col>": <png_bytes> } }
    """
    results: dict[str, dict[str, bytes]] = {}
    for sheet_name, df in data.items():
        sheet_imgs: dict[str, bytes] = {}
        for r_idx, row in enumerate(df.iter_rows(named=False)):
            for c_idx, val in enumerate(row):
                text = str(val) if val is not None else ""
                if _has_math(text):
                    png = render_math_png(text)
                    if png:
                        sheet_imgs[f"R{r_idx}C{c_idx}"] = png
        if sheet_imgs:
            results[sheet_name] = sheet_imgs
    return results


def extract_docx_equations(path: str) -> list[dict]:
    """
    Extract OMML (Office Math Markup Language) equations from a DOCX file.
    Returns list of { "text": str, "png": bytes | None }.
    """
    results: list[dict] = []
    try:
        from lxml import etree
        from zipfile import ZipFile

        NS_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
        NS_M = "http://schemas.openxmlformats.org/officeDocument/2006/math"

        with ZipFile(path) as zf:
            if "word/document.xml" not in zf.namelist():
                return results
            xml_bytes = zf.read("word/document.xml")

        root = etree.fromstring(xml_bytes)

        for omath in root.iter(f"{{{NS_M}}}oMath"):
            # Best-effort text extraction from OMML
            text_parts = []
            for t_elem in omath.iter(f"{{{NS_M}}}t"):
                if t_elem.text:
                    text_parts.append(t_elem.text)
            # Also grab any plain <w:t> inside runs inside the math
            for t_elem in omath.iter(f"{{{NS_W}}}t"):
                if t_elem.text:
                    text_parts.append(t_elem.text)

            raw_text = " ".join(text_parts).strip()
            if not raw_text:
                continue

            # Try to render as LaTeX/math expression
            png = render_math_png(raw_text) if raw_text else None
            results.append({"text": raw_text, "png": png})

    except Exception as e:
        logger.debug(f"DOCX equation extraction failed: {e}")
    return results


def extract_pptx_equations(path: str) -> list[dict]:
    """
    Extract math equations from a PPTX file.
    Returns list of { "slide": int, "text": str, "png": bytes | None }.
    """
    results: list[dict] = []
    try:
        from lxml import etree
        from zipfile import ZipFile

        NS_A14 = "http://schemas.microsoft.com/office/drawing/2010/main"
        NS_M   = "http://schemas.openxmlformats.org/officeDocument/2006/math"

        with ZipFile(path) as zf:
            slide_files = sorted(
                n for n in zf.namelist()
                if n.startswith("ppt/slides/slide") and n.endswith(".xml")
            )
            for slide_idx, slide_file in enumerate(slide_files, 1):
                xml_bytes = zf.read(slide_file)
                root = etree.fromstring(xml_bytes)
                for omath in root.iter(f"{{{NS_M}}}oMath"):
                    text_parts = [
                        t.text for t in omath.iter(f"{{{NS_M}}}t")
                        if t.text
                    ]
                    raw_text = " ".join(text_parts).strip()
                    if raw_text:
                        png = render_math_png(raw_text)
                        results.append({
                            "slide": slide_idx,
                            "text": raw_text,
                            "png": png,
                        })
    except Exception as e:
        logger.debug(f"PPTX equation extraction failed: {e}")
    return results


def enrich_pdf_with_charts(
    doc,  # fitz.Document (already built, will get new pages appended)
    charts: list[dict],
    title_prefix: str = "المخططات البيانية",
    equations: list[dict] | None = None,
) -> None:
    """
    Append a new page (or pages) to a PyMuPDF document containing
    chart images and optional equation images.
    """
    if not charts and not equations:
        return

    try:
        import pymupdf as fitz
        FONT = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
        PAGE_W, PAGE_H = 595, 842  # A4 portrait
        MARGIN = 30
        TITLE_H = 24
        GAP = 12

        page = doc.new_page(width=PAGE_W, height=PAGE_H)
        try:
            page.insert_font(fontname="dvb", fontfile=FONT)
        except Exception:
            pass

        y = MARGIN

        # Section title
        if charts:
            page.draw_rect(
                fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + TITLE_H),
                color=None, fill=(79/255, 70/255, 229/255), overlay=True,
            )
            page.insert_textbox(
                fitz.Rect(MARGIN + 6, y + 4, PAGE_W - MARGIN - 6, y + TITLE_H - 4),
                title_prefix,
                fontname="dvb", fontsize=11, color=(1, 1, 1),
                align=1,  # center
                overlay=True,
            )
            y += TITLE_H + GAP

        for ch in charts:
            png = ch.get("png")
            if not png:
                continue
            ch_title = ch.get("title", "")

            # If we're near the bottom, start a new page
            if y + 180 > PAGE_H - MARGIN:
                page = doc.new_page(width=PAGE_W, height=PAGE_H)
                try:
                    page.insert_font(fontname="dvb", fontfile=FONT)
                except Exception:
                    pass
                y = MARGIN

            # Chart title label
            if ch_title:
                page.insert_textbox(
                    fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + 14),
                    ch_title,
                    fontname="dvb", fontsize=9, color=(0.2, 0.2, 0.2),
                    align=1, overlay=True,
                )
                y += 16

            # Embed chart image
            available_w = PAGE_W - 2 * MARGIN
            img_stream = fitz.open(stream=png, filetype="png")
            img_w = img_stream[0].rect.width
            img_h = img_stream[0].rect.height
            img_stream.close()

            scale = min(available_w / max(img_w, 1), 200 / max(img_h, 1))
            draw_w = img_w * scale
            draw_h = img_h * scale
            x0 = MARGIN + (available_w - draw_w) / 2

            rect = fitz.Rect(x0, y, x0 + draw_w, y + draw_h)
            page.insert_image(rect, stream=png, overlay=True)
            y += draw_h + GAP

        # Equations section
        if equations:
            if y + TITLE_H + 20 > PAGE_H - MARGIN:
                page = doc.new_page(width=PAGE_W, height=PAGE_H)
                try:
                    page.insert_font(fontname="dvb", fontfile=FONT)
                except Exception:
                    pass
                y = MARGIN

            page.draw_rect(
                fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + TITLE_H),
                color=None, fill=(30/255, 100/255, 60/255), overlay=True,
            )
            page.insert_textbox(
                fitz.Rect(MARGIN + 6, y + 4, PAGE_W - MARGIN - 6, y + TITLE_H - 4),
                "المعادلات والصيغ الرياضية",
                fontname="dvb", fontsize=11, color=(1, 1, 1),
                align=1, overlay=True,
            )
            y += TITLE_H + GAP

            for eq in equations:
                png = eq.get("png")
                text = eq.get("text", "")
                if not png and not text:
                    continue

                if y + 50 > PAGE_H - MARGIN:
                    page = doc.new_page(width=PAGE_W, height=PAGE_H)
                    try:
                        page.insert_font(fontname="dvb", fontfile=FONT)
                    except Exception:
                        pass
                    y = MARGIN

                if png:
                    img_stream = fitz.open(stream=png, filetype="png")
                    img_w = img_stream[0].rect.width
                    img_h = img_stream[0].rect.height
                    img_stream.close()
                    available_w = PAGE_W - 2 * MARGIN
                    scale = min(available_w / max(img_w, 1), 60 / max(img_h, 1))
                    draw_w = img_w * scale
                    draw_h = img_h * scale
                    x0 = MARGIN + (available_w - draw_w) / 2
                    rect = fitz.Rect(x0, y, x0 + draw_w, y + draw_h)
                    page.insert_image(rect, stream=png, overlay=True)
                    y += draw_h + 4
                else:
                    page.insert_textbox(
                        fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + 18),
                        text,
                        fontname="dvb", fontsize=8, color=(0.1, 0.1, 0.5),
                        align=0, overlay=True,
                    )
                    y += 20

    except Exception as e:
        logger.warning(f"enrich_pdf_with_charts failed: {e}")


def charts_to_html_blocks(charts: list[dict], equations: list[dict] | None = None) -> str:
    """
    Convert extracted charts and equations to HTML <img> tags (base64 PNG).
    Returns an HTML string fragment ready to inject into an HTML page.
    """
    import base64
    parts: list[str] = []

    if charts:
        parts.append(
            '<div style="margin:24px 0;"><h3 style="font-family:sans-serif;'
            'color:#4F46E5;border-bottom:2px solid #4F46E5;padding-bottom:6px;">'
            '📊 المخططات البيانية</h3>'
        )
        for ch in charts:
            png = ch.get("png")
            title = ch.get("title", "")
            if not png:
                continue
            b64 = base64.b64encode(png).decode()
            parts.append(
                f'<figure style="margin:16px 0;text-align:center;">'
                f'<img src="data:image/png;base64,{b64}" '
                f'style="max-width:100%;border:1px solid #e2e8f0;border-radius:8px;'
                f'box-shadow:0 2px 8px rgba(0,0,0,.08);" alt="{title}"/>'
                f'{"<figcaption style=\"font-family:sans-serif;font-size:13px;color:#64748b;margin-top:6px;\">" + title + "</figcaption>" if title else ""}'
                f'</figure>'
            )
        parts.append('</div>')

    if equations:
        parts.append(
            '<div style="margin:24px 0;"><h3 style="font-family:sans-serif;'
            'color:#059669;border-bottom:2px solid #059669;padding-bottom:6px;">'
            '∑ المعادلات والصيغ الرياضية</h3>'
        )
        for eq in equations:
            png = eq.get("png")
            text = eq.get("text", "")
            if png:
                b64 = base64.b64encode(png).decode()
                parts.append(
                    f'<figure style="margin:12px 0;text-align:center;">'
                    f'<img src="data:image/png;base64,{b64}" '
                    f'style="max-height:60px;" alt="{text}"/>'
                    f'</figure>'
                )
            elif text:
                parts.append(
                    f'<div style="font-family:monospace;font-size:13px;'
                    f'background:#f1f5f9;padding:8px 12px;border-radius:6px;'
                    f'margin:8px 0;">{text}</div>'
                )
        parts.append('</div>')

    return "\n".join(parts)

_LATEX_CMDS    = re.compile(
    r'\\(?:'
    # ── Core math ──────────────────────────────────────────────────────────
    r'frac|dfrac|tfrac|cfrac|sqrt|sum|int|iint|iiint|oint|prod|lim|'
    r'limsup|liminf|sup|inf|max|min|arg|det|dim|ker|'
    # ── Greek letters (lowercase) ───────────────────────────────────────────
    r'alpha|beta|gamma|delta|epsilon|varepsilon|zeta|eta|theta|vartheta|'
    r'iota|kappa|lambda|mu|nu|xi|pi|varpi|rho|varrho|sigma|varsigma|'
    r'tau|upsilon|phi|varphi|chi|psi|omega|'
    # ── Greek letters (uppercase) ───────────────────────────────────────────
    r'Gamma|Delta|Theta|Lambda|Xi|Pi|Sigma|Upsilon|Phi|Psi|Omega|'
    # ── Physics symbols ─────────────────────────────────────────────────────
    r'hbar|ell|nabla|partial|infty|angle|measuredangle|sphericalangle|'
    r'perp|parallel|propto|therefore|because|Re|Im|'
    r'vec|hat|tilde|bar|dot|ddot|dddot|ddddot|'
    r'overrightarrow|overleftarrow|overleftrightarrow|'
    # ── Operators & relations ───────────────────────────────────────────────
    r'pm|mp|times|div|cdot|circ|bullet|star|ast|oplus|ominus|otimes|oslash|'
    r'leq|geq|neq|ll|gg|approx|equiv|sim|simeq|cong|'
    r'subset|supset|subseteq|supseteq|in|notin|ni|'
    r'cup|cap|setminus|emptyset|varnothing|'
    r'wedge|vee|neg|forall|exists|nexists|'
    # ── Arrows ──────────────────────────────────────────────────────────────
    r'to|gets|rightarrow|leftarrow|Rightarrow|Leftarrow|'
    r'leftrightarrow|Leftrightarrow|iff|mapsto|'
    r'uparrow|downarrow|Uparrow|Downarrow|updownarrow|'
    r'nearrow|searrow|nwarrow|swarrow|'
    # ── Trig / calculus / analysis ──────────────────────────────────────────
    r'sin|cos|tan|cot|sec|csc|arcsin|arccos|arctan|'
    r'sinh|cosh|tanh|coth|log|ln|exp|'
    # ── Brackets & delimiters ───────────────────────────────────────────────
    r'left|right|lfloor|rfloor|lceil|rceil|langle|rangle|'
    r'bigl|bigr|Bigl|Bigr|biggl|biggr|'
    # ── Environments / structure ────────────────────────────────────────────
    r'begin|end|'
    r'matrix|bmatrix|pmatrix|vmatrix|Vmatrix|Bmatrix|'
    r'cases|array|aligned|align|gather|equation|'
    # ── Text / style ────────────────────────────────────────────────────────
    r'text|mbox|mathrm|mathbf|mathbb|mathcal|mathit|mathsf|mathtt|'
    r'boldsymbol|bm|'
    # ── Over/under decorations ──────────────────────────────────────────────
    r'overline|underline|overbrace|underbrace|widehat|widetilde|'
    r'overrightarrow|xleftarrow|xrightarrow|'
    # ── Combinatorics / algebra ─────────────────────────────────────────────
    r'binom|choose|pmod|mod|gcd|lcm|'
    # ── Chemistry (mhchem-style plain text) ────────────────────────────────
    r'ce|chem|chemfig|'
    # ── Engineering / signals ───────────────────────────────────────────────
    r'laplace|fourier|ztransform|'
    r'degree|celsius|fahrenheit|ohm|'
    # ── Misc ────────────────────────────────────────────────────────────────
    r'quad|qquad|,|;|:|!|'
    r'ldots|cdots|vdots|ddots|'
    r'prime|dagger|ddagger|'
    r'infty|aleph|beth|'
    r'hline|vline|cline'
    r')'
)

# Chemical formula pattern: sequences like H2O, CO2, C6H12O6, Fe2O3,
# NaCl, CaCO3, H2SO4 — must have at least one digit or be multi-element
_CHEM_FORMULA  = re.compile(
    r'\b(?:[A-Z][a-z]?\d*){2,}\b|'      # multi-element: NaCl, H2O, Fe2O3
    r'\b[A-Z][a-z]?\d+\b'               # single element with count: H2, O3
)


def _has_math(text: str) -> bool:
    """Quick check: does this string look like it contains math/LaTeX?"""
    if not text:
        return False
    return bool(
        _LATEX_DISPLAY.search(text) or
        _LATEX_INLINE.search(text) or
        _LATEX_CMDS.search(text)
    )


def _strip_latex_delimiters(expr: str) -> str:
    """Remove surrounding $ or $$ delimiters."""
    expr = expr.strip()
    if expr.startswith("$$") and expr.endswith("$$"):
        return expr[2:-2].strip()
    if expr.startswith("$") and expr.endswith("$"):
        return expr[1:-1].strip()
    return expr


# ── Matplotlib helpers ────────────────────────────────────────────────────────

def _mpl_backend() -> None:
    """Ensure a non-interactive Agg backend is active."""
    import matplotlib
    matplotlib.use("Agg")


def render_math_png(expr: str, fontsize: int = 14, dpi: int = 150) -> bytes | None:
    """
    Render a mathematical expression (LaTeX) to PNG bytes.
    Returns None if rendering fails.
    """
    try:
        _mpl_backend()
        import matplotlib.pyplot as plt

        clean = _strip_latex_delimiters(expr)
        # Wrap in $…$ if not already a display environment
        if not (clean.startswith(r"\begin") or clean.startswith(r"\[")):
            latex = f"${clean}$"
        else:
            latex = clean

        fig = plt.figure(figsize=(6, 1.2))
        fig.patch.set_facecolor("white")
        t = fig.text(
            0.5, 0.5, latex,
            ha="center", va="center",
            fontsize=fontsize,
            color="black",
        )
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=dpi,
                    bbox_inches="tight", pad_inches=0.15,
                    facecolor="white")
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        logger.debug(f"Math render failed for '{expr[:40]}': {e}")
        return None


# ── Excel chart extraction ────────────────────────────────────────────────────

def _resolve_ref(ws, ref) -> list:
    """Resolve an openpyxl Reference object to a flat list of values."""
    if ref is None:
        return []
    vals: list = []
    try:
        for row in ws.iter_rows(
            min_row=ref.min_row, max_row=ref.max_row,
            min_col=ref.min_col, max_col=ref.max_col,
            values_only=True,
        ):
            for v in row:
                if v is not None:
                    vals.append(v)
    except Exception:
        pass
    return vals


def _chart_title(chart) -> str:
    """Best-effort chart title extraction."""
    try:
        t = chart.title
        if t is None:
            return ""
        if isinstance(t, str):
            return t
        # openpyxl Title object
        if hasattr(t, "tx") and t.tx:
            tx = t.tx
            if hasattr(tx, "rich") and tx.rich:
                parts = []
                for para in tx.rich.paragraphs:
                    for run in para.runs:
                        if run.t:
                            parts.append(run.t)
                return "".join(parts)
            if hasattr(tx, "strRef") and tx.strRef and tx.strRef.f:
                return tx.strRef.f
        return ""
    except Exception:
        return ""


def _series_values(ws, series) -> tuple[list, list, str]:
    """
    Return (categories, values, series_label) from an openpyxl series object.
    Works for BarChart, LineChart, AreaChart, PieChart series.
    """
    cats: list = []
    vals: list = []
    label = ""
    try:
        # Series title / legend entry
        if hasattr(series, "title") and series.title:
            st = series.title
            if hasattr(st, "v") and st.v:
                label = str(st.v)
            elif hasattr(st, "strRef") and st.strRef:
                f = getattr(st.strRef, "f", "")
                label = f.split("!")[-1].strip('"') if "!" in f else f
        # Values
        val_ref = getattr(series, "val", None) or getattr(series, "yVal", None)
        if val_ref and hasattr(val_ref, "numRef") and val_ref.numRef:
            nr = val_ref.numRef
            if hasattr(nr, "ref") and nr.ref:
                from openpyxl.utils import range_to_tuple
                try:
                    ws2 = ws.parent[nr.ref.split("!")[0].strip("'")] \
                        if "!" in nr.ref else ws
                    ref_str = nr.ref.split("!")[-1] if "!" in nr.ref else nr.ref
                    from openpyxl import load_workbook as _lw
                    # Use cached values if available
                    if nr.numCache and nr.numCache.pt:
                        vals = [float(p.v) for p in nr.numCache.pt if p.v is not None]
                    else:
                        from openpyxl.utils.cell import range_boundaries
                        mn_col, mn_row, mx_col, mx_row = range_boundaries(ref_str)
                        for row in ws.iter_rows(
                            min_row=mn_row, max_row=mx_row,
                            min_col=mn_col, max_col=mx_col,
                            values_only=True,
                        ):
                            for v in row:
                                if isinstance(v, (int, float)):
                                    vals.append(float(v))
                except Exception:
                    pass
        # Categories
        cat_ref = getattr(series, "cat", None) or getattr(series, "xVal", None)
        if cat_ref:
            if hasattr(cat_ref, "numRef") and cat_ref.numRef and cat_ref.numRef.numCache:
                cats = [str(p.v) for p in cat_ref.numRef.numCache.pt if p.v is not None]
            elif hasattr(cat_ref, "strRef") and cat_ref.strRef and cat_ref.strRef.strCache:
                cats = [str(p.v) for p in cat_ref.strRef.strCache.pt if p.v is not None]
    except Exception as e:
        logger.debug(f"Series extraction error: {e}")
    return cats, vals, label


def _render_chart(chart, ws) -> bytes | None:
    """Render a single openpyxl chart object to PNG bytes using matplotlib."""
    try:
        _mpl_backend()
        import matplotlib.pyplot as plt
        import matplotlib.cm as cm
        import numpy as np

        chart_type = type(chart).__name__.lower()  # barchart, linechart, piechart …
        title = _chart_title(chart)
        series_list = list(chart.series)

        if not series_list:
            return None

        fig, ax = plt.subplots(figsize=(7, 4))
        fig.patch.set_facecolor("white")
        ax.set_facecolor("#f8f9fa")
        if title:
            ax.set_title(title, fontsize=11, fontweight="bold", pad=8)

        colors = cm.tab10.colors

        if "pie" in chart_type:
            cats, vals, _ = _series_values(ws, series_list[0])
            if not vals:
                plt.close(fig)
                return None
            labels = cats if cats else [f"Item {i+1}" for i in range(len(vals))]
            ax.pie(vals, labels=labels, autopct="%1.1f%%",
                   colors=colors[:len(vals)], startangle=90)

        elif "scatter" in chart_type or "bubble" in chart_type:
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                x = list(range(len(vals))) if not cats else cats
                try:
                    x = [float(v) for v in x]
                except (ValueError, TypeError):
                    x = list(range(len(vals)))
                if vals:
                    ax.scatter(x, vals, label=label or f"Series {idx+1}",
                               color=colors[idx % len(colors)], s=60, alpha=0.8)
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        elif "line" in chart_type or "area" in chart_type or "stock" in chart_type:
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                x = list(range(len(vals))) if not cats else cats
                if vals:
                    if "area" in chart_type:
                        ax.fill_between(range(len(vals)), vals,
                                        alpha=0.4, color=colors[idx % len(colors)])
                    ax.plot(vals, label=label or f"Series {idx+1}",
                            color=colors[idx % len(colors)], linewidth=2, marker="o",
                            markersize=4)
                    if cats:
                        step = max(1, len(cats) // 10)
                        ax.set_xticks(range(0, len(cats), step))
                        ax.set_xticklabels(
                            [str(cats[i]) for i in range(0, len(cats), step)],
                            rotation=30, fontsize=7, ha="right"
                        )
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        elif "radar" in chart_type:
            cats, vals, _ = _series_values(ws, series_list[0])
            if vals and cats:
                angles = np.linspace(0, 2 * np.pi, len(vals), endpoint=False).tolist()
                angles += angles[:1]
                vals_closed = vals + vals[:1]
                ax_r = fig.add_subplot(111, polar=True)
                ax.remove()
                ax_r.plot(angles, vals_closed, linewidth=1.5, color=colors[0])
                ax_r.fill(angles, vals_closed, alpha=0.25, color=colors[0])
                ax_r.set_xticks(angles[:-1])
                ax_r.set_xticklabels(cats, fontsize=8)

        else:
            # Default: grouped bar chart
            all_cats: list[str] = []
            all_vals: list[tuple[list, str]] = []
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                if vals:
                    all_vals.append((vals, label or f"Series {idx+1}"))
                    if cats and not all_cats:
                        all_cats = [str(c) for c in cats]

            if not all_vals:
                plt.close(fig)
                return None

            n_groups = max(len(v) for v, _ in all_vals)
            n_series = len(all_vals)
            x = np.arange(n_groups)
            width = 0.8 / n_series

            for idx, (vals, label) in enumerate(all_vals):
                offset = (idx - n_series / 2 + 0.5) * width
                bars = ax.bar(x[:len(vals)] + offset, vals, width,
                              label=label, color=colors[idx % len(colors)], alpha=0.85)

            if all_cats:
                step = max(1, len(all_cats) // 12)
                ax.set_xticks(x[::step])
                ax.set_xticklabels(
                    [all_cats[i] for i in range(0, len(all_cats), step)],
                    rotation=30, fontsize=7, ha="right",
                )
            if n_series > 1:
                ax.legend(fontsize=8)
            ax.grid(axis="y", alpha=0.3)

        plt.tight_layout(pad=0.8)
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=130, bbox_inches="tight",
                    facecolor="white")
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        logger.warning(f"Chart render failed [{type(chart).__name__}]: {e}")
        try:
            plt.close("all")
        except Exception:
            pass
        return None


# ── Public API ────────────────────────────────────────────────────────────────

def extract_excel_charts(path: str, sheet_name: str | None = None) -> list[dict]:
    """
    Extract all charts from an Excel worksheet.

    Returns a list of dicts:
        {
          "title": str,
          "png":   bytes,      # rendered PNG image
        }
    Returns an empty list on any failure.
    """
    results: list[dict] = []
    try:
        import openpyxl
        # Load with data_only=True so cached values are available for reference resolution
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet_names = wb.sheetnames if sheet_name is None else [sheet_name]
        for sn in sheet_names:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            charts = getattr(ws, "_charts", [])
            for chart in charts:
                png = _render_chart(chart, ws)
                if png:
                    results.append({
                        "title": _chart_title(chart) or sn,
                        "png": png,
                    })
        wb.close()
    except Exception as e:
        logger.warning(f"Excel chart extraction failed: {e}")
    return results


def extract_excel_formulas(path: str, sheet_name: str | None = None) -> dict[str, dict]:
    """
    Read formula strings from Excel cells (loads WITHOUT data_only).

    Returns:
        { sheet_name: { "A1": "=SUM(B1:B10)", ... } }
    """
    result: dict[str, dict] = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=False)
        names = wb.sheetnames if sheet_name is None else [sheet_name]
        for sn in names:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            sheet_formulas: dict[str, str] = {}
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        sheet_formulas[cell.coordinate] = v
            if sheet_formulas:
                result[sn] = sheet_formulas
        wb.close()
    except Exception as e:
        logger.debug(f"Formula extraction failed: {e}")
    return result


def detect_math_cells(data: dict[str, Any]) -> dict[str, dict[str, bytes]]:
    """
    Scan a dict of {sheet_name: DataFrame} and detect cells with LaTeX/math content.

    Returns:
        { sheet_name: { "R<row>C<col>": <png_bytes> } }
    """
    results: dict[str, dict[str, bytes]] = {}
    for sheet_name, df in data.items():
        sheet_imgs: dict[str, bytes] = {}
        for r_idx, row in enumerate(df.iter_rows(named=False)):
            for c_idx, val in enumerate(row):
                text = str(val) if val is not None else ""
                if _has_math(text):
                    png = render_math_png(text)
                    if png:
                        sheet_imgs[f"R{r_idx}C{c_idx}"] = png
        if sheet_imgs:
            results[sheet_name] = sheet_imgs
    return results


def extract_docx_equations(path: str) -> list[dict]:
    """
    Extract OMML (Office Math Markup Language) equations from a DOCX file.
    Returns list of { "text": str, "png": bytes | None }.
    """
    results: list[dict] = []
    try:
        from lxml import etree
        from zipfile import ZipFile

        NS_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
        NS_M = "http://schemas.openxmlformats.org/officeDocument/2006/math"

        with ZipFile(path) as zf:
            if "word/document.xml" not in zf.namelist():
                return results
            xml_bytes = zf.read("word/document.xml")

        root = etree.fromstring(xml_bytes)

        for omath in root.iter(f"{{{NS_M}}}oMath"):
            # Best-effort text extraction from OMML
            text_parts = []
            for t_elem in omath.iter(f"{{{NS_M}}}t"):
                if t_elem.text:
                    text_parts.append(t_elem.text)
            # Also grab any plain <w:t> inside runs inside the math
            for t_elem in omath.iter(f"{{{NS_W}}}t"):
                if t_elem.text:
                    text_parts.append(t_elem.text)

            raw_text = " ".join(text_parts).strip()
            if not raw_text:
                continue

            # Try to render as LaTeX/math expression
            png = render_math_png(raw_text) if raw_text else None
            results.append({"text": raw_text, "png": png})

    except Exception as e:
        logger.debug(f"DOCX equation extraction failed: {e}")
    return results


def extract_pptx_equations(path: str) -> list[dict]:
    """
    Extract math equations from a PPTX file.
    Returns list of { "slide": int, "text": str, "png": bytes | None }.
    """
    results: list[dict] = []
    try:
        from lxml import etree
        from zipfile import ZipFile

        NS_A14 = "http://schemas.microsoft.com/office/drawing/2010/main"
        NS_M   = "http://schemas.openxmlformats.org/officeDocument/2006/math"

        with ZipFile(path) as zf:
            slide_files = sorted(
                n for n in zf.namelist()
                if n.startswith("ppt/slides/slide") and n.endswith(".xml")
            )
            for slide_idx, slide_file in enumerate(slide_files, 1):
                xml_bytes = zf.read(slide_file)
                root = etree.fromstring(xml_bytes)
                for omath in root.iter(f"{{{NS_M}}}oMath"):
                    text_parts = [
                        t.text for t in omath.iter(f"{{{NS_M}}}t")
                        if t.text
                    ]
                    raw_text = " ".join(text_parts).strip()
                    if raw_text:
                        png = render_math_png(raw_text)
                        results.append({
                            "slide": slide_idx,
                            "text": raw_text,
                            "png": png,
                        })
    except Exception as e:
        logger.debug(f"PPTX equation extraction failed: {e}")
    return results


def enrich_pdf_with_charts(
    doc,  # fitz.Document (already built, will get new pages appended)
    charts: list[dict],
    title_prefix: str = "المخططات البيانية",
    equations: list[dict] | None = None,
) -> None:
    """
    Append a new page (or pages) to a PyMuPDF document containing
    chart images and optional equation images.
    """
    if not charts and not equations:
        return

    try:
        import pymupdf as fitz
        FONT = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
        PAGE_W, PAGE_H = 595, 842  # A4 portrait
        MARGIN = 30
        TITLE_H = 24
        GAP = 12

        page = doc.new_page(width=PAGE_W, height=PAGE_H)
        try:
            page.insert_font(fontname="dvb", fontfile=FONT)
        except Exception:
            pass

        y = MARGIN

        # Section title
        if charts:
            page.draw_rect(
                fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + TITLE_H),
                color=None, fill=(79/255, 70/255, 229/255), overlay=True,
            )
            page.insert_textbox(
                fitz.Rect(MARGIN + 6, y + 4, PAGE_W - MARGIN - 6, y + TITLE_H - 4),
                title_prefix,
                fontname="dvb", fontsize=11, color=(1, 1, 1),
                align=1,  # center
                overlay=True,
            )
            y += TITLE_H + GAP

        for ch in charts:
            png = ch.get("png")
            if not png:
                continue
            ch_title = ch.get("title", "")

            # If we're near the bottom, start a new page
            if y + 180 > PAGE_H - MARGIN:
                page = doc.new_page(width=PAGE_W, height=PAGE_H)
                try:
                    page.insert_font(fontname="dvb", fontfile=FONT)
                except Exception:
                    pass
                y = MARGIN

            # Chart title label
            if ch_title:
                page.insert_textbox(
                    fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + 14),
                    ch_title,
                    fontname="dvb", fontsize=9, color=(0.2, 0.2, 0.2),
                    align=1, overlay=True,
                )
                y += 16

            # Embed chart image
            available_w = PAGE_W - 2 * MARGIN
            img_stream = fitz.open(stream=png, filetype="png")
            img_w = img_stream[0].rect.width
            img_h = img_stream[0].rect.height
            img_stream.close()

            scale = min(available_w / max(img_w, 1), 200 / max(img_h, 1))
            draw_w = img_w * scale
            draw_h = img_h * scale
            x0 = MARGIN + (available_w - draw_w) / 2

            rect = fitz.Rect(x0, y, x0 + draw_w, y + draw_h)
            page.insert_image(rect, stream=png, overlay=True)
            y += draw_h + GAP

        # Equations section
        if equations:
            if y + TITLE_H + 20 > PAGE_H - MARGIN:
                page = doc.new_page(width=PAGE_W, height=PAGE_H)
                try:
                    page.insert_font(fontname="dvb", fontfile=FONT)
                except Exception:
                    pass
                y = MARGIN

            page.draw_rect(
                fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + TITLE_H),
                color=None, fill=(30/255, 100/255, 60/255), overlay=True,
            )
            page.insert_textbox(
                fitz.Rect(MARGIN + 6, y + 4, PAGE_W - MARGIN - 6, y + TITLE_H - 4),
                "المعادلات والصيغ الرياضية",
                fontname="dvb", fontsize=11, color=(1, 1, 1),
                align=1, overlay=True,
            )
            y += TITLE_H + GAP

            for eq in equations:
                png = eq.get("png")
                text = eq.get("text", "")
                if not png and not text:
                    continue

                if y + 50 > PAGE_H - MARGIN:
                    page = doc.new_page(width=PAGE_W, height=PAGE_H)
                    try:
                        page.insert_font(fontname="dvb", fontfile=FONT)
                    except Exception:
                        pass
                    y = MARGIN

                if png:
                    img_stream = fitz.open(stream=png, filetype="png")
                    img_w = img_stream[0].rect.width
                    img_h = img_stream[0].rect.height
                    img_stream.close()
                    available_w = PAGE_W - 2 * MARGIN
                    scale = min(available_w / max(img_w, 1), 60 / max(img_h, 1))
                    draw_w = img_w * scale
                    draw_h = img_h * scale
                    x0 = MARGIN + (available_w - draw_w) / 2
                    rect = fitz.Rect(x0, y, x0 + draw_w, y + draw_h)
                    page.insert_image(rect, stream=png, overlay=True)
                    y += draw_h + 4
                else:
                    page.insert_textbox(
                        fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + 18),
                        text,
                        fontname="dvb", fontsize=8, color=(0.1, 0.1, 0.5),
                        align=0, overlay=True,
                    )
                    y += 20

    except Exception as e:
        logger.warning(f"enrich_pdf_with_charts failed: {e}")


def charts_to_html_blocks(charts: list[dict], equations: list[dict] | None = None) -> str:
    """
    Convert extracted charts and equations to HTML <img> tags (base64 PNG).
    Returns an HTML string fragment ready to inject into an HTML page.
    """
    import base64
    parts: list[str] = []

    if charts:
        parts.append(
            '<div style="margin:24px 0;"><h3 style="font-family:sans-serif;'
            'color:#4F46E5;border-bottom:2px solid #4F46E5;padding-bottom:6px;">'
            '📊 المخططات البيانية</h3>'
        )
        for ch in charts:
            png = ch.get("png")
            title = ch.get("title", "")
            if not png:
                continue
            b64 = base64.b64encode(png).decode()
            parts.append(
                f'<figure style="margin:16px 0;text-align:center;">'
                f'<img src="data:image/png;base64,{b64}" '
                f'style="max-width:100%;border:1px solid #e2e8f0;border-radius:8px;'
                f'box-shadow:0 2px 8px rgba(0,0,0,.08);" alt="{title}"/>'
                f'{"<figcaption style=\"font-family:sans-serif;font-size:13px;color:#64748b;margin-top:6px;\">" + title + "</figcaption>" if title else ""}'
                f'</figure>'
            )
        parts.append('</div>')

    if equations:
        parts.append(
            '<div style="margin:24px 0;"><h3 style="font-family:sans-serif;'
            'color:#059669;border-bottom:2px solid #059669;padding-bottom:6px;">'
            '∑ المعادلات والصيغ الرياضية</h3>'
        )
        for eq in equations:
            png = eq.get("png")
            text = eq.get("text", "")
            if png:
                b64 = base64.b64encode(png).decode()
                parts.append(
                    f'<figure style="margin:12px 0;text-align:center;">'
                    f'<img src="data:image/png;base64,{b64}" '
                    f'style="max-height:60px;" alt="{text}"/>'
                    f'</figure>'
                )
            elif text:
                parts.append(
                    f'<div style="font-family:monospace;font-size:13px;'
                    f'background:#f1f5f9;padding:8px 12px;border-radius:6px;'
                    f'margin:8px 0;">{text}</div>'
                )
        parts.append('</div>')

    return "\n".join(parts)


def _has_math(text: str) -> bool:
    """Quick check: does this string look like it contains math/LaTeX?"""
    if not text:
        return False
    return bool(
        _LATEX_DISPLAY.search(text) or
        _LATEX_INLINE.search(text) or
        _LATEX_CMDS.search(text)
    )


def _strip_latex_delimiters(expr: str) -> str:
    """Remove surrounding $ or $$ delimiters."""
    expr = expr.strip()
    if expr.startswith("$$") and expr.endswith("$$"):
        return expr[2:-2].strip()
    if expr.startswith("$") and expr.endswith("$"):
        return expr[1:-1].strip()
    return expr


# ── Matplotlib helpers ────────────────────────────────────────────────────────

def _mpl_backend() -> None:
    """Ensure a non-interactive Agg backend is active."""
    import matplotlib
    matplotlib.use("Agg")


def render_math_png(expr: str, fontsize: int = 14, dpi: int = 150) -> bytes | None:
    """
    Render a mathematical expression (LaTeX) to PNG bytes.
    Returns None if rendering fails.
    """
    try:
        _mpl_backend()
        import matplotlib.pyplot as plt

        clean = _strip_latex_delimiters(expr)
        # Wrap in $…$ if not already a display environment
        if not (clean.startswith(r"\begin") or clean.startswith(r"\[")):
            latex = f"${clean}$"
        else:
            latex = clean

        fig = plt.figure(figsize=(6, 1.2))
        fig.patch.set_facecolor("white")
        t = fig.text(
            0.5, 0.5, latex,
            ha="center", va="center",
            fontsize=fontsize,
            color="black",
        )
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=dpi,
                    bbox_inches="tight", pad_inches=0.15,
                    facecolor="white")
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        logger.debug(f"Math render failed for '{expr[:40]}': {e}")
        return None


# ── Excel chart extraction ────────────────────────────────────────────────────

def _resolve_ref(ws, ref) -> list:
    """Resolve an openpyxl Reference object to a flat list of values."""
    if ref is None:
        return []
    vals: list = []
    try:
        for row in ws.iter_rows(
            min_row=ref.min_row, max_row=ref.max_row,
            min_col=ref.min_col, max_col=ref.max_col,
            values_only=True,
        ):
            for v in row:
                if v is not None:
                    vals.append(v)
    except Exception:
        pass
    return vals


def _chart_title(chart) -> str:
    """Best-effort chart title extraction."""
    try:
        t = chart.title
        if t is None:
            return ""
        if isinstance(t, str):
            return t
        # openpyxl Title object
        if hasattr(t, "tx") and t.tx:
            tx = t.tx
            if hasattr(tx, "rich") and tx.rich:
                parts = []
                for para in tx.rich.paragraphs:
                    for run in para.runs:
                        if run.t:
                            parts.append(run.t)
                return "".join(parts)
            if hasattr(tx, "strRef") and tx.strRef and tx.strRef.f:
                return tx.strRef.f
        return ""
    except Exception:
        return ""


def _series_values(ws, series) -> tuple[list, list, str]:
    """
    Return (categories, values, series_label) from an openpyxl series object.
    Works for BarChart, LineChart, AreaChart, PieChart series.
    """
    cats: list = []
    vals: list = []
    label = ""
    try:
        # Series title / legend entry
        if hasattr(series, "title") and series.title:
            st = series.title
            if hasattr(st, "v") and st.v:
                label = str(st.v)
            elif hasattr(st, "strRef") and st.strRef:
                f = getattr(st.strRef, "f", "")
                label = f.split("!")[-1].strip('"') if "!" in f else f
        # Values
        val_ref = getattr(series, "val", None) or getattr(series, "yVal", None)
        if val_ref and hasattr(val_ref, "numRef") and val_ref.numRef:
            nr = val_ref.numRef
            if hasattr(nr, "ref") and nr.ref:
                from openpyxl.utils import range_to_tuple
                try:
                    ws2 = ws.parent[nr.ref.split("!")[0].strip("'")] \
                        if "!" in nr.ref else ws
                    ref_str = nr.ref.split("!")[-1] if "!" in nr.ref else nr.ref
                    from openpyxl import load_workbook as _lw
                    # Use cached values if available
                    if nr.numCache and nr.numCache.pt:
                        vals = [float(p.v) for p in nr.numCache.pt if p.v is not None]
                    else:
                        from openpyxl.utils.cell import range_boundaries
                        mn_col, mn_row, mx_col, mx_row = range_boundaries(ref_str)
                        for row in ws.iter_rows(
                            min_row=mn_row, max_row=mx_row,
                            min_col=mn_col, max_col=mx_col,
                            values_only=True,
                        ):
                            for v in row:
                                if isinstance(v, (int, float)):
                                    vals.append(float(v))
                except Exception:
                    pass
        # Categories
        cat_ref = getattr(series, "cat", None) or getattr(series, "xVal", None)
        if cat_ref:
            if hasattr(cat_ref, "numRef") and cat_ref.numRef and cat_ref.numRef.numCache:
                cats = [str(p.v) for p in cat_ref.numRef.numCache.pt if p.v is not None]
            elif hasattr(cat_ref, "strRef") and cat_ref.strRef and cat_ref.strRef.strCache:
                cats = [str(p.v) for p in cat_ref.strRef.strCache.pt if p.v is not None]
    except Exception as e:
        logger.debug(f"Series extraction error: {e}")
    return cats, vals, label


def _render_chart(chart, ws) -> bytes | None:
    """Render a single openpyxl chart object to PNG bytes using matplotlib."""
    try:
        _mpl_backend()
        import matplotlib.pyplot as plt
        import matplotlib.cm as cm
        import numpy as np

        chart_type = type(chart).__name__.lower()  # barchart, linechart, piechart …
        title = _chart_title(chart)
        series_list = list(chart.series)

        if not series_list:
            return None

        fig, ax = plt.subplots(figsize=(7, 4))
        fig.patch.set_facecolor("white")
        ax.set_facecolor("#f8f9fa")
        if title:
            ax.set_title(title, fontsize=11, fontweight="bold", pad=8)

        colors = cm.tab10.colors

        if "pie" in chart_type:
            cats, vals, _ = _series_values(ws, series_list[0])
            if not vals:
                plt.close(fig)
                return None
            labels = cats if cats else [f"Item {i+1}" for i in range(len(vals))]
            ax.pie(vals, labels=labels, autopct="%1.1f%%",
                   colors=colors[:len(vals)], startangle=90)

        elif "scatter" in chart_type or "bubble" in chart_type:
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                x = list(range(len(vals))) if not cats else cats
                try:
                    x = [float(v) for v in x]
                except (ValueError, TypeError):
                    x = list(range(len(vals)))
                if vals:
                    ax.scatter(x, vals, label=label or f"Series {idx+1}",
                               color=colors[idx % len(colors)], s=60, alpha=0.8)
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        elif "line" in chart_type or "area" in chart_type or "stock" in chart_type:
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                x = list(range(len(vals))) if not cats else cats
                if vals:
                    if "area" in chart_type:
                        ax.fill_between(range(len(vals)), vals,
                                        alpha=0.4, color=colors[idx % len(colors)])
                    ax.plot(vals, label=label or f"Series {idx+1}",
                            color=colors[idx % len(colors)], linewidth=2, marker="o",
                            markersize=4)
                    if cats:
                        step = max(1, len(cats) // 10)
                        ax.set_xticks(range(0, len(cats), step))
                        ax.set_xticklabels(
                            [str(cats[i]) for i in range(0, len(cats), step)],
                            rotation=30, fontsize=7, ha="right"
                        )
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        elif "radar" in chart_type:
            cats, vals, _ = _series_values(ws, series_list[0])
            if vals and cats:
                angles = np.linspace(0, 2 * np.pi, len(vals), endpoint=False).tolist()
                angles += angles[:1]
                vals_closed = vals + vals[:1]
                ax_r = fig.add_subplot(111, polar=True)
                ax.remove()
                ax_r.plot(angles, vals_closed, linewidth=1.5, color=colors[0])
                ax_r.fill(angles, vals_closed, alpha=0.25, color=colors[0])
                ax_r.set_xticks(angles[:-1])
                ax_r.set_xticklabels(cats, fontsize=8)

        else:
            # Default: grouped bar chart
            all_cats: list[str] = []
            all_vals: list[tuple[list, str]] = []
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                if vals:
                    all_vals.append((vals, label or f"Series {idx+1}"))
                    if cats and not all_cats:
                        all_cats = [str(c) for c in cats]

            if not all_vals:
                plt.close(fig)
                return None

            n_groups = max(len(v) for v, _ in all_vals)
            n_series = len(all_vals)
            x = np.arange(n_groups)
            width = 0.8 / n_series

            for idx, (vals, label) in enumerate(all_vals):
                offset = (idx - n_series / 2 + 0.5) * width
                bars = ax.bar(x[:len(vals)] + offset, vals, width,
                              label=label, color=colors[idx % len(colors)], alpha=0.85)

            if all_cats:
                step = max(1, len(all_cats) // 12)
                ax.set_xticks(x[::step])
                ax.set_xticklabels(
                    [all_cats[i] for i in range(0, len(all_cats), step)],
                    rotation=30, fontsize=7, ha="right",
                )
            if n_series > 1:
                ax.legend(fontsize=8)
            ax.grid(axis="y", alpha=0.3)

        plt.tight_layout(pad=0.8)
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=130, bbox_inches="tight",
                    facecolor="white")
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        logger.warning(f"Chart render failed [{type(chart).__name__}]: {e}")
        try:
            plt.close("all")
        except Exception:
            pass
        return None


# ── Public API ────────────────────────────────────────────────────────────────

def extract_excel_charts(path: str, sheet_name: str | None = None) -> list[dict]:
    """
    Extract all charts from an Excel worksheet.

    Returns a list of dicts:
        {
          "title": str,
          "png":   bytes,      # rendered PNG image
        }
    Returns an empty list on any failure.
    """
    results: list[dict] = []
    try:
        import openpyxl
        # Load with data_only=True so cached values are available for reference resolution
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet_names = wb.sheetnames if sheet_name is None else [sheet_name]
        for sn in sheet_names:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            charts = getattr(ws, "_charts", [])
            for chart in charts:
                png = _render_chart(chart, ws)
                if png:
                    results.append({
                        "title": _chart_title(chart) or sn,
                        "png": png,
                    })
        wb.close()
    except Exception as e:
        logger.warning(f"Excel chart extraction failed: {e}")
    return results


def extract_excel_formulas(path: str, sheet_name: str | None = None) -> dict[str, dict]:
    """
    Read formula strings from Excel cells (loads WITHOUT data_only).

    Returns:
        { sheet_name: { "A1": "=SUM(B1:B10)", ... } }
    """
    result: dict[str, dict] = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=False)
        names = wb.sheetnames if sheet_name is None else [sheet_name]
        for sn in names:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            sheet_formulas: dict[str, str] = {}
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        sheet_formulas[cell.coordinate] = v
            if sheet_formulas:
                result[sn] = sheet_formulas
        wb.close()
    except Exception as e:
        logger.debug(f"Formula extraction failed: {e}")
    return result


def detect_math_cells(data: dict[str, Any]) -> dict[str, dict[str, bytes]]:
    """
    Scan a dict of {sheet_name: DataFrame} and detect cells with LaTeX/math content.

    Returns:
        { sheet_name: { "R<row>C<col>": <png_bytes> } }
    """
    results: dict[str, dict[str, bytes]] = {}
    for sheet_name, df in data.items():
        sheet_imgs: dict[str, bytes] = {}
        for r_idx, row in enumerate(df.iter_rows(named=False)):
            for c_idx, val in enumerate(row):
                text = str(val) if val is not None else ""
                if _has_math(text):
                    png = render_math_png(text)
                    if png:
                        sheet_imgs[f"R{r_idx}C{c_idx}"] = png
        if sheet_imgs:
            results[sheet_name] = sheet_imgs
    return results


def extract_docx_equations(path: str) -> list[dict]:
    """
    Extract OMML (Office Math Markup Language) equations from a DOCX file.
    Returns list of { "text": str, "png": bytes | None }.
    """
    results: list[dict] = []
    try:
        from lxml import etree
        from zipfile import ZipFile

        NS_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
        NS_M = "http://schemas.openxmlformats.org/officeDocument/2006/math"

        with ZipFile(path) as zf:
            if "word/document.xml" not in zf.namelist():
                return results
            xml_bytes = zf.read("word/document.xml")

        root = etree.fromstring(xml_bytes)

        for omath in root.iter(f"{{{NS_M}}}oMath"):
            # Best-effort text extraction from OMML
            text_parts = []
            for t_elem in omath.iter(f"{{{NS_M}}}t"):
                if t_elem.text:
                    text_parts.append(t_elem.text)
            # Also grab any plain <w:t> inside runs inside the math
            for t_elem in omath.iter(f"{{{NS_W}}}t"):
                if t_elem.text:
                    text_parts.append(t_elem.text)

            raw_text = " ".join(text_parts).strip()
            if not raw_text:
                continue

            # Try to render as LaTeX/math expression
            png = render_math_png(raw_text) if raw_text else None
            results.append({"text": raw_text, "png": png})

    except Exception as e:
        logger.debug(f"DOCX equation extraction failed: {e}")
    return results


def extract_pptx_equations(path: str) -> list[dict]:
    """
    Extract math equations from a PPTX file.
    Returns list of { "slide": int, "text": str, "png": bytes | None }.
    """
    results: list[dict] = []
    try:
        from lxml import etree
        from zipfile import ZipFile

        NS_A14 = "http://schemas.microsoft.com/office/drawing/2010/main"
        NS_M   = "http://schemas.openxmlformats.org/officeDocument/2006/math"

        with ZipFile(path) as zf:
            slide_files = sorted(
                n for n in zf.namelist()
                if n.startswith("ppt/slides/slide") and n.endswith(".xml")
            )
            for slide_idx, slide_file in enumerate(slide_files, 1):
                xml_bytes = zf.read(slide_file)
                root = etree.fromstring(xml_bytes)
                for omath in root.iter(f"{{{NS_M}}}oMath"):
                    text_parts = [
                        t.text for t in omath.iter(f"{{{NS_M}}}t")
                        if t.text
                    ]
                    raw_text = " ".join(text_parts).strip()
                    if raw_text:
                        png = render_math_png(raw_text)
                        results.append({
                            "slide": slide_idx,
                            "text": raw_text,
                            "png": png,
                        })
    except Exception as e:
        logger.debug(f"PPTX equation extraction failed: {e}")
    return results


def enrich_pdf_with_charts(
    doc,  # fitz.Document (already built, will get new pages appended)
    charts: list[dict],
    title_prefix: str = "المخططات البيانية",
    equations: list[dict] | None = None,
) -> None:
    """
    Append a new page (or pages) to a PyMuPDF document containing
    chart images and optional equation images.
    """
    if not charts and not equations:
        return

    try:
        import pymupdf as fitz
        FONT = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
        PAGE_W, PAGE_H = 595, 842  # A4 portrait
        MARGIN = 30
        TITLE_H = 24
        GAP = 12

        page = doc.new_page(width=PAGE_W, height=PAGE_H)
        try:
            page.insert_font(fontname="dvb", fontfile=FONT)
        except Exception:
            pass

        y = MARGIN

        # Section title
        if charts:
            page.draw_rect(
                fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + TITLE_H),
                color=None, fill=(79/255, 70/255, 229/255), overlay=True,
            )
            page.insert_textbox(
                fitz.Rect(MARGIN + 6, y + 4, PAGE_W - MARGIN - 6, y + TITLE_H - 4),
                title_prefix,
                fontname="dvb", fontsize=11, color=(1, 1, 1),
                align=1,  # center
                overlay=True,
            )
            y += TITLE_H + GAP

        for ch in charts:
            png = ch.get("png")
            if not png:
                continue
            ch_title = ch.get("title", "")

            # If we're near the bottom, start a new page
            if y + 180 > PAGE_H - MARGIN:
                page = doc.new_page(width=PAGE_W, height=PAGE_H)
                try:
                    page.insert_font(fontname="dvb", fontfile=FONT)
                except Exception:
                    pass
                y = MARGIN

            # Chart title label
            if ch_title:
                page.insert_textbox(
                    fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + 14),
                    ch_title,
                    fontname="dvb", fontsize=9, color=(0.2, 0.2, 0.2),
                    align=1, overlay=True,
                )
                y += 16

            # Embed chart image
            available_w = PAGE_W - 2 * MARGIN
            img_stream = fitz.open(stream=png, filetype="png")
            img_w = img_stream[0].rect.width
            img_h = img_stream[0].rect.height
            img_stream.close()

            scale = min(available_w / max(img_w, 1), 200 / max(img_h, 1))
            draw_w = img_w * scale
            draw_h = img_h * scale
            x0 = MARGIN + (available_w - draw_w) / 2

            rect = fitz.Rect(x0, y, x0 + draw_w, y + draw_h)
            page.insert_image(rect, stream=png, overlay=True)
            y += draw_h + GAP

        # Equations section
        if equations:
            if y + TITLE_H + 20 > PAGE_H - MARGIN:
                page = doc.new_page(width=PAGE_W, height=PAGE_H)
                try:
                    page.insert_font(fontname="dvb", fontfile=FONT)
                except Exception:
                    pass
                y = MARGIN

            page.draw_rect(
                fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + TITLE_H),
                color=None, fill=(30/255, 100/255, 60/255), overlay=True,
            )
            page.insert_textbox(
                fitz.Rect(MARGIN + 6, y + 4, PAGE_W - MARGIN - 6, y + TITLE_H - 4),
                "المعادلات والصيغ الرياضية",
                fontname="dvb", fontsize=11, color=(1, 1, 1),
                align=1, overlay=True,
            )
            y += TITLE_H + GAP

            for eq in equations:
                png = eq.get("png")
                text = eq.get("text", "")
                if not png and not text:
                    continue

                if y + 50 > PAGE_H - MARGIN:
                    page = doc.new_page(width=PAGE_W, height=PAGE_H)
                    try:
                        page.insert_font(fontname="dvb", fontfile=FONT)
                    except Exception:
                        pass
                    y = MARGIN

                if png:
                    img_stream = fitz.open(stream=png, filetype="png")
                    img_w = img_stream[0].rect.width
                    img_h = img_stream[0].rect.height
                    img_stream.close()
                    available_w = PAGE_W - 2 * MARGIN
                    scale = min(available_w / max(img_w, 1), 60 / max(img_h, 1))
                    draw_w = img_w * scale
                    draw_h = img_h * scale
                    x0 = MARGIN + (available_w - draw_w) / 2
                    rect = fitz.Rect(x0, y, x0 + draw_w, y + draw_h)
                    page.insert_image(rect, stream=png, overlay=True)
                    y += draw_h + 4
                else:
                    page.insert_textbox(
                        fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + 18),
                        text,
                        fontname="dvb", fontsize=8, color=(0.1, 0.1, 0.5),
                        align=0, overlay=True,
                    )
                    y += 20

    except Exception as e:
        logger.warning(f"enrich_pdf_with_charts failed: {e}")


def charts_to_html_blocks(charts: list[dict], equations: list[dict] | None = None) -> str:
    """
    Convert extracted charts and equations to HTML <img> tags (base64 PNG).
    Returns an HTML string fragment ready to inject into an HTML page.
    """
    import base64
    parts: list[str] = []

    if charts:
        parts.append(
            '<div style="margin:24px 0;"><h3 style="font-family:sans-serif;'
            'color:#4F46E5;border-bottom:2px solid #4F46E5;padding-bottom:6px;">'
            '📊 المخططات البيانية</h3>'
        )
        for ch in charts:
            png = ch.get("png")
            title = ch.get("title", "")
            if not png:
                continue
            b64 = base64.b64encode(png).decode()
            parts.append(
                f'<figure style="margin:16px 0;text-align:center;">'
                f'<img src="data:image/png;base64,{b64}" '
                f'style="max-width:100%;border:1px solid #e2e8f0;border-radius:8px;'
                f'box-shadow:0 2px 8px rgba(0,0,0,.08);" alt="{title}"/>'
                f'{"<figcaption style=\"font-family:sans-serif;font-size:13px;color:#64748b;margin-top:6px;\">" + title + "</figcaption>" if title else ""}'
                f'</figure>'
            )
        parts.append('</div>')

    if equations:
        parts.append(
            '<div style="margin:24px 0;"><h3 style="font-family:sans-serif;'
            'color:#059669;border-bottom:2px solid #059669;padding-bottom:6px;">'
            '∑ المعادلات والصيغ الرياضية</h3>'
        )
        for eq in equations:
            png = eq.get("png")
            text = eq.get("text", "")
            if png:
                b64 = base64.b64encode(png).decode()
                parts.append(
                    f'<figure style="margin:12px 0;text-align:center;">'
                    f'<img src="data:image/png;base64,{b64}" '
                    f'style="max-height:60px;" alt="{text}"/>'
                    f'</figure>'
                )
            elif text:
                parts.append(
                    f'<div style="font-family:monospace;font-size:13px;'
                    f'background:#f1f5f9;padding:8px 12px;border-radius:6px;'
                    f'margin:8px 0;">{text}</div>'
                )
        parts.append('</div>')

    return "\n".join(parts)
_LATEX_DISPLAY = re.compile(r'\$\$.+?\$\$', re.DOTALL)


def _has_math(text: str) -> bool:
    """Quick check: does this string look like it contains math/LaTeX?"""
    if not text:
        return False
    return bool(
        _LATEX_DISPLAY.search(text) or
        _LATEX_INLINE.search(text) or
        _LATEX_CMDS.search(text)
    )


def _strip_latex_delimiters(expr: str) -> str:
    """Remove surrounding $ or $$ delimiters."""
    expr = expr.strip()
    if expr.startswith("$$") and expr.endswith("$$"):
        return expr[2:-2].strip()
    if expr.startswith("$") and expr.endswith("$"):
        return expr[1:-1].strip()
    return expr


# ── Matplotlib helpers ────────────────────────────────────────────────────────

def _mpl_backend() -> None:
    """Ensure a non-interactive Agg backend is active."""
    import matplotlib
    matplotlib.use("Agg")


def render_math_png(expr: str, fontsize: int = 14, dpi: int = 150) -> bytes | None:
    """
    Render a mathematical expression (LaTeX) to PNG bytes.
    Returns None if rendering fails.
    """
    try:
        _mpl_backend()
        import matplotlib.pyplot as plt

        clean = _strip_latex_delimiters(expr)
        # Wrap in $…$ if not already a display environment
        if not (clean.startswith(r"\begin") or clean.startswith(r"\[")):
            latex = f"${clean}$"
        else:
            latex = clean

        fig = plt.figure(figsize=(6, 1.2))
        fig.patch.set_facecolor("white")
        t = fig.text(
            0.5, 0.5, latex,
            ha="center", va="center",
            fontsize=fontsize,
            color="black",
        )
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=dpi,
                    bbox_inches="tight", pad_inches=0.15,
                    facecolor="white")
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        logger.debug(f"Math render failed for '{expr[:40]}': {e}")
        return None


# ── Excel chart extraction ────────────────────────────────────────────────────

def _resolve_ref(ws, ref) -> list:
    """Resolve an openpyxl Reference object to a flat list of values."""
    if ref is None:
        return []
    vals: list = []
    try:
        for row in ws.iter_rows(
            min_row=ref.min_row, max_row=ref.max_row,
            min_col=ref.min_col, max_col=ref.max_col,
            values_only=True,
        ):
            for v in row:
                if v is not None:
                    vals.append(v)
    except Exception:
        pass
    return vals


def _chart_title(chart) -> str:
    """Best-effort chart title extraction."""
    try:
        t = chart.title
        if t is None:
            return ""
        if isinstance(t, str):
            return t
        # openpyxl Title object
        if hasattr(t, "tx") and t.tx:
            tx = t.tx
            if hasattr(tx, "rich") and tx.rich:
                parts = []
                for para in tx.rich.paragraphs:
                    for run in para.runs:
                        if run.t:
                            parts.append(run.t)
                return "".join(parts)
            if hasattr(tx, "strRef") and tx.strRef and tx.strRef.f:
                return tx.strRef.f
        return ""
    except Exception:
        return ""


def _series_values(ws, series) -> tuple[list, list, str]:
    """
    Return (categories, values, series_label) from an openpyxl series object.
    Works for BarChart, LineChart, AreaChart, PieChart series.
    """
    cats: list = []
    vals: list = []
    label = ""
    try:
        # Series title / legend entry
        if hasattr(series, "title") and series.title:
            st = series.title
            if hasattr(st, "v") and st.v:
                label = str(st.v)
            elif hasattr(st, "strRef") and st.strRef:
                f = getattr(st.strRef, "f", "")
                label = f.split("!")[-1].strip('"') if "!" in f else f
        # Values
        val_ref = getattr(series, "val", None) or getattr(series, "yVal", None)
        if val_ref and hasattr(val_ref, "numRef") and val_ref.numRef:
            nr = val_ref.numRef
            if hasattr(nr, "ref") and nr.ref:
                from openpyxl.utils import range_to_tuple
                try:
                    ws2 = ws.parent[nr.ref.split("!")[0].strip("'")] \
                        if "!" in nr.ref else ws
                    ref_str = nr.ref.split("!")[-1] if "!" in nr.ref else nr.ref
                    from openpyxl import load_workbook as _lw
                    # Use cached values if available
                    if nr.numCache and nr.numCache.pt:
                        vals = [float(p.v) for p in nr.numCache.pt if p.v is not None]
                    else:
                        from openpyxl.utils.cell import range_boundaries
                        mn_col, mn_row, mx_col, mx_row = range_boundaries(ref_str)
                        for row in ws.iter_rows(
                            min_row=mn_row, max_row=mx_row,
                            min_col=mn_col, max_col=mx_col,
                            values_only=True,
                        ):
                            for v in row:
                                if isinstance(v, (int, float)):
                                    vals.append(float(v))
                except Exception:
                    pass
        # Categories
        cat_ref = getattr(series, "cat", None) or getattr(series, "xVal", None)
        if cat_ref:
            if hasattr(cat_ref, "numRef") and cat_ref.numRef and cat_ref.numRef.numCache:
                cats = [str(p.v) for p in cat_ref.numRef.numCache.pt if p.v is not None]
            elif hasattr(cat_ref, "strRef") and cat_ref.strRef and cat_ref.strRef.strCache:
                cats = [str(p.v) for p in cat_ref.strRef.strCache.pt if p.v is not None]
    except Exception as e:
        logger.debug(f"Series extraction error: {e}")
    return cats, vals, label


def _render_chart(chart, ws) -> bytes | None:
    """Render a single openpyxl chart object to PNG bytes using matplotlib."""
    try:
        _mpl_backend()
        import matplotlib.pyplot as plt
        import matplotlib.cm as cm
        import numpy as np

        chart_type = type(chart).__name__.lower()  # barchart, linechart, piechart …
        title = _chart_title(chart)
        series_list = list(chart.series)

        if not series_list:
            return None

        fig, ax = plt.subplots(figsize=(7, 4))
        fig.patch.set_facecolor("white")
        ax.set_facecolor("#f8f9fa")
        if title:
            ax.set_title(title, fontsize=11, fontweight="bold", pad=8)

        colors = cm.tab10.colors

        if "pie" in chart_type:
            cats, vals, _ = _series_values(ws, series_list[0])
            if not vals:
                plt.close(fig)
                return None
            labels = cats if cats else [f"Item {i+1}" for i in range(len(vals))]
            ax.pie(vals, labels=labels, autopct="%1.1f%%",
                   colors=colors[:len(vals)], startangle=90)

        elif "scatter" in chart_type or "bubble" in chart_type:
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                x = list(range(len(vals))) if not cats else cats
                try:
                    x = [float(v) for v in x]
                except (ValueError, TypeError):
                    x = list(range(len(vals)))
                if vals:
                    ax.scatter(x, vals, label=label or f"Series {idx+1}",
                               color=colors[idx % len(colors)], s=60, alpha=0.8)
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        elif "line" in chart_type or "area" in chart_type or "stock" in chart_type:
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                x = list(range(len(vals))) if not cats else cats
                if vals:
                    if "area" in chart_type:
                        ax.fill_between(range(len(vals)), vals,
                                        alpha=0.4, color=colors[idx % len(colors)])
                    ax.plot(vals, label=label or f"Series {idx+1}",
                            color=colors[idx % len(colors)], linewidth=2, marker="o",
                            markersize=4)
                    if cats:
                        step = max(1, len(cats) // 10)
                        ax.set_xticks(range(0, len(cats), step))
                        ax.set_xticklabels(
                            [str(cats[i]) for i in range(0, len(cats), step)],
                            rotation=30, fontsize=7, ha="right"
                        )
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        elif "radar" in chart_type:
            cats, vals, _ = _series_values(ws, series_list[0])
            if vals and cats:
                angles = np.linspace(0, 2 * np.pi, len(vals), endpoint=False).tolist()
                angles += angles[:1]
                vals_closed = vals + vals[:1]
                ax_r = fig.add_subplot(111, polar=True)
                ax.remove()
                ax_r.plot(angles, vals_closed, linewidth=1.5, color=colors[0])
                ax_r.fill(angles, vals_closed, alpha=0.25, color=colors[0])
                ax_r.set_xticks(angles[:-1])
                ax_r.set_xticklabels(cats, fontsize=8)

        else:
            # Default: grouped bar chart
            all_cats: list[str] = []
            all_vals: list[tuple[list, str]] = []
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                if vals:
                    all_vals.append((vals, label or f"Series {idx+1}"))
                    if cats and not all_cats:
                        all_cats = [str(c) for c in cats]

            if not all_vals:
                plt.close(fig)
                return None

            n_groups = max(len(v) for v, _ in all_vals)
            n_series = len(all_vals)
            x = np.arange(n_groups)
            width = 0.8 / n_series

            for idx, (vals, label) in enumerate(all_vals):
                offset = (idx - n_series / 2 + 0.5) * width
                bars = ax.bar(x[:len(vals)] + offset, vals, width,
                              label=label, color=colors[idx % len(colors)], alpha=0.85)

            if all_cats:
                step = max(1, len(all_cats) // 12)
                ax.set_xticks(x[::step])
                ax.set_xticklabels(
                    [all_cats[i] for i in range(0, len(all_cats), step)],
                    rotation=30, fontsize=7, ha="right",
                )
            if n_series > 1:
                ax.legend(fontsize=8)
            ax.grid(axis="y", alpha=0.3)

        plt.tight_layout(pad=0.8)
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=130, bbox_inches="tight",
                    facecolor="white")
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        logger.warning(f"Chart render failed [{type(chart).__name__}]: {e}")
        try:
            plt.close("all")
        except Exception:
            pass
        return None


# ── Public API ────────────────────────────────────────────────────────────────

def extract_excel_charts(path: str, sheet_name: str | None = None) -> list[dict]:
    """
    Extract all charts from an Excel worksheet.

    Returns a list of dicts:
        {
          "title": str,
          "png":   bytes,      # rendered PNG image
        }
    Returns an empty list on any failure.
    """
    results: list[dict] = []
    try:
        import openpyxl
        # Load with data_only=True so cached values are available for reference resolution
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet_names = wb.sheetnames if sheet_name is None else [sheet_name]
        for sn in sheet_names:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            charts = getattr(ws, "_charts", [])
            for chart in charts:
                png = _render_chart(chart, ws)
                if png:
                    results.append({
                        "title": _chart_title(chart) or sn,
                        "png": png,
                    })
        wb.close()
    except Exception as e:
        logger.warning(f"Excel chart extraction failed: {e}")
    return results


def extract_excel_formulas(path: str, sheet_name: str | None = None) -> dict[str, dict]:
    """
    Read formula strings from Excel cells (loads WITHOUT data_only).

    Returns:
        { sheet_name: { "A1": "=SUM(B1:B10)", ... } }
    """
    result: dict[str, dict] = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=False)
        names = wb.sheetnames if sheet_name is None else [sheet_name]
        for sn in names:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            sheet_formulas: dict[str, str] = {}
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        sheet_formulas[cell.coordinate] = v
            if sheet_formulas:
                result[sn] = sheet_formulas
        wb.close()
    except Exception as e:
        logger.debug(f"Formula extraction failed: {e}")
    return result


def detect_math_cells(data: dict[str, Any]) -> dict[str, dict[str, bytes]]:
    """
    Scan a dict of {sheet_name: DataFrame} and detect cells with LaTeX/math content.

    Returns:
        { sheet_name: { "R<row>C<col>": <png_bytes> } }
    """
    results: dict[str, dict[str, bytes]] = {}
    for sheet_name, df in data.items():
        sheet_imgs: dict[str, bytes] = {}
        for r_idx, row in enumerate(df.iter_rows(named=False)):
            for c_idx, val in enumerate(row):
                text = str(val) if val is not None else ""
                if _has_math(text):
                    png = render_math_png(text)
                    if png:
                        sheet_imgs[f"R{r_idx}C{c_idx}"] = png
        if sheet_imgs:
            results[sheet_name] = sheet_imgs
    return results


def extract_docx_equations(path: str) -> list[dict]:
    """
    Extract OMML (Office Math Markup Language) equations from a DOCX file.
    Returns list of { "text": str, "png": bytes | None }.
    """
    results: list[dict] = []
    try:
        from lxml import etree
        from zipfile import ZipFile

        NS_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
        NS_M = "http://schemas.openxmlformats.org/officeDocument/2006/math"

        with ZipFile(path) as zf:
            if "word/document.xml" not in zf.namelist():
                return results
            xml_bytes = zf.read("word/document.xml")

        root = etree.fromstring(xml_bytes)

        for omath in root.iter(f"{{{NS_M}}}oMath"):
            # Best-effort text extraction from OMML
            text_parts = []
            for t_elem in omath.iter(f"{{{NS_M}}}t"):
                if t_elem.text:
                    text_parts.append(t_elem.text)
            # Also grab any plain <w:t> inside runs inside the math
            for t_elem in omath.iter(f"{{{NS_W}}}t"):
                if t_elem.text:
                    text_parts.append(t_elem.text)

            raw_text = " ".join(text_parts).strip()
            if not raw_text:
                continue

            # Try to render as LaTeX/math expression
            png = render_math_png(raw_text) if raw_text else None
            results.append({"text": raw_text, "png": png})

    except Exception as e:
        logger.debug(f"DOCX equation extraction failed: {e}")
    return results


def extract_pptx_equations(path: str) -> list[dict]:
    """
    Extract math equations from a PPTX file.
    Returns list of { "slide": int, "text": str, "png": bytes | None }.
    """
    results: list[dict] = []
    try:
        from lxml import etree
        from zipfile import ZipFile

        NS_A14 = "http://schemas.microsoft.com/office/drawing/2010/main"
        NS_M   = "http://schemas.openxmlformats.org/officeDocument/2006/math"

        with ZipFile(path) as zf:
            slide_files = sorted(
                n for n in zf.namelist()
                if n.startswith("ppt/slides/slide") and n.endswith(".xml")
            )
            for slide_idx, slide_file in enumerate(slide_files, 1):
                xml_bytes = zf.read(slide_file)
                root = etree.fromstring(xml_bytes)
                for omath in root.iter(f"{{{NS_M}}}oMath"):
                    text_parts = [
                        t.text for t in omath.iter(f"{{{NS_M}}}t")
                        if t.text
                    ]
                    raw_text = " ".join(text_parts).strip()
                    if raw_text:
                        png = render_math_png(raw_text)
                        results.append({
                            "slide": slide_idx,
                            "text": raw_text,
                            "png": png,
                        })
    except Exception as e:
        logger.debug(f"PPTX equation extraction failed: {e}")
    return results


def enrich_pdf_with_charts(
    doc,  # fitz.Document (already built, will get new pages appended)
    charts: list[dict],
    title_prefix: str = "المخططات البيانية",
    equations: list[dict] | None = None,
) -> None:
    """
    Append a new page (or pages) to a PyMuPDF document containing
    chart images and optional equation images.
    """
    if not charts and not equations:
        return

    try:
        import pymupdf as fitz
        FONT = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
        PAGE_W, PAGE_H = 595, 842  # A4 portrait
        MARGIN = 30
        TITLE_H = 24
        GAP = 12

        page = doc.new_page(width=PAGE_W, height=PAGE_H)
        try:
            page.insert_font(fontname="dvb", fontfile=FONT)
        except Exception:
            pass

        y = MARGIN

        # Section title
        if charts:
            page.draw_rect(
                fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + TITLE_H),
                color=None, fill=(79/255, 70/255, 229/255), overlay=True,
            )
            page.insert_textbox(
                fitz.Rect(MARGIN + 6, y + 4, PAGE_W - MARGIN - 6, y + TITLE_H - 4),
                title_prefix,
                fontname="dvb", fontsize=11, color=(1, 1, 1),
                align=1,  # center
                overlay=True,
            )
            y += TITLE_H + GAP

        for ch in charts:
            png = ch.get("png")
            if not png:
                continue
            ch_title = ch.get("title", "")

            # If we're near the bottom, start a new page
            if y + 180 > PAGE_H - MARGIN:
                page = doc.new_page(width=PAGE_W, height=PAGE_H)
                try:
                    page.insert_font(fontname="dvb", fontfile=FONT)
                except Exception:
                    pass
                y = MARGIN

            # Chart title label
            if ch_title:
                page.insert_textbox(
                    fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + 14),
                    ch_title,
                    fontname="dvb", fontsize=9, color=(0.2, 0.2, 0.2),
                    align=1, overlay=True,
                )
                y += 16

            # Embed chart image
            available_w = PAGE_W - 2 * MARGIN
            img_stream = fitz.open(stream=png, filetype="png")
            img_w = img_stream[0].rect.width
            img_h = img_stream[0].rect.height
            img_stream.close()

            scale = min(available_w / max(img_w, 1), 200 / max(img_h, 1))
            draw_w = img_w * scale
            draw_h = img_h * scale
            x0 = MARGIN + (available_w - draw_w) / 2

            rect = fitz.Rect(x0, y, x0 + draw_w, y + draw_h)
            page.insert_image(rect, stream=png, overlay=True)
            y += draw_h + GAP

        # Equations section
        if equations:
            if y + TITLE_H + 20 > PAGE_H - MARGIN:
                page = doc.new_page(width=PAGE_W, height=PAGE_H)
                try:
                    page.insert_font(fontname="dvb", fontfile=FONT)
                except Exception:
                    pass
                y = MARGIN

            page.draw_rect(
                fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + TITLE_H),
                color=None, fill=(30/255, 100/255, 60/255), overlay=True,
            )
            page.insert_textbox(
                fitz.Rect(MARGIN + 6, y + 4, PAGE_W - MARGIN - 6, y + TITLE_H - 4),
                "المعادلات والصيغ الرياضية",
                fontname="dvb", fontsize=11, color=(1, 1, 1),
                align=1, overlay=True,
            )
            y += TITLE_H + GAP

            for eq in equations:
                png = eq.get("png")
                text = eq.get("text", "")
                if not png and not text:
                    continue

                if y + 50 > PAGE_H - MARGIN:
                    page = doc.new_page(width=PAGE_W, height=PAGE_H)
                    try:
                        page.insert_font(fontname="dvb", fontfile=FONT)
                    except Exception:
                        pass
                    y = MARGIN

                if png:
                    img_stream = fitz.open(stream=png, filetype="png")
                    img_w = img_stream[0].rect.width
                    img_h = img_stream[0].rect.height
                    img_stream.close()
                    available_w = PAGE_W - 2 * MARGIN
                    scale = min(available_w / max(img_w, 1), 60 / max(img_h, 1))
                    draw_w = img_w * scale
                    draw_h = img_h * scale
                    x0 = MARGIN + (available_w - draw_w) / 2
                    rect = fitz.Rect(x0, y, x0 + draw_w, y + draw_h)
                    page.insert_image(rect, stream=png, overlay=True)
                    y += draw_h + 4
                else:
                    page.insert_textbox(
                        fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + 18),
                        text,
                        fontname="dvb", fontsize=8, color=(0.1, 0.1, 0.5),
                        align=0, overlay=True,
                    )
                    y += 20

    except Exception as e:
        logger.warning(f"enrich_pdf_with_charts failed: {e}")


def charts_to_html_blocks(charts: list[dict], equations: list[dict] | None = None) -> str:
    """
    Convert extracted charts and equations to HTML <img> tags (base64 PNG).
    Returns an HTML string fragment ready to inject into an HTML page.
    """
    import base64
    parts: list[str] = []

    if charts:
        parts.append(
            '<div style="margin:24px 0;"><h3 style="font-family:sans-serif;'
            'color:#4F46E5;border-bottom:2px solid #4F46E5;padding-bottom:6px;">'
            '📊 المخططات البيانية</h3>'
        )
        for ch in charts:
            png = ch.get("png")
            title = ch.get("title", "")
            if not png:
                continue
            b64 = base64.b64encode(png).decode()
            parts.append(
                f'<figure style="margin:16px 0;text-align:center;">'
                f'<img src="data:image/png;base64,{b64}" '
                f'style="max-width:100%;border:1px solid #e2e8f0;border-radius:8px;'
                f'box-shadow:0 2px 8px rgba(0,0,0,.08);" alt="{title}"/>'
                f'{"<figcaption style=\"font-family:sans-serif;font-size:13px;color:#64748b;margin-top:6px;\">" + title + "</figcaption>" if title else ""}'
                f'</figure>'
            )
        parts.append('</div>')

    if equations:
        parts.append(
            '<div style="margin:24px 0;"><h3 style="font-family:sans-serif;'
            'color:#059669;border-bottom:2px solid #059669;padding-bottom:6px;">'
            '∑ المعادلات والصيغ الرياضية</h3>'
        )
        for eq in equations:
            png = eq.get("png")
            text = eq.get("text", "")
            if png:
                b64 = base64.b64encode(png).decode()
                parts.append(
                    f'<figure style="margin:12px 0;text-align:center;">'
                    f'<img src="data:image/png;base64,{b64}" '
                    f'style="max-height:60px;" alt="{text}"/>'
                    f'</figure>'
                )
            elif text:
                parts.append(
                    f'<div style="font-family:monospace;font-size:13px;'
                    f'background:#f1f5f9;padding:8px 12px;border-radius:6px;'
                    f'margin:8px 0;">{text}</div>'
                )
        parts.append('</div>')

    return "\n".join(parts)
_LATEX_CMDS    = re.compile(
    r'\\(?:'
    # ── Core math ──────────────────────────────────────────────────────────
    r'frac|dfrac|tfrac|cfrac|sqrt|sum|int|iint|iiint|oint|prod|lim|'
    r'limsup|liminf|sup|inf|max|min|arg|det|dim|ker|'
    # ── Greek letters (lowercase) ───────────────────────────────────────────
    r'alpha|beta|gamma|delta|epsilon|varepsilon|zeta|eta|theta|vartheta|'
    r'iota|kappa|lambda|mu|nu|xi|pi|varpi|rho|varrho|sigma|varsigma|'
    r'tau|upsilon|phi|varphi|chi|psi|omega|'
    # ── Greek letters (uppercase) ───────────────────────────────────────────
    r'Gamma|Delta|Theta|Lambda|Xi|Pi|Sigma|Upsilon|Phi|Psi|Omega|'
    # ── Physics symbols ─────────────────────────────────────────────────────
    r'hbar|ell|nabla|partial|infty|angle|measuredangle|sphericalangle|'
    r'perp|parallel|propto|therefore|because|Re|Im|'
    r'vec|hat|tilde|bar|dot|ddot|dddot|ddddot|'
    r'overrightarrow|overleftarrow|overleftrightarrow|'
    # ── Operators & relations ───────────────────────────────────────────────
    r'pm|mp|times|div|cdot|circ|bullet|star|ast|oplus|ominus|otimes|oslash|'
    r'leq|geq|neq|ll|gg|approx|equiv|sim|simeq|cong|'
    r'subset|supset|subseteq|supseteq|in|notin|ni|'
    r'cup|cap|setminus|emptyset|varnothing|'
    r'wedge|vee|neg|forall|exists|nexists|'
    # ── Arrows ──────────────────────────────────────────────────────────────
    r'to|gets|rightarrow|leftarrow|Rightarrow|Leftarrow|'
    r'leftrightarrow|Leftrightarrow|iff|mapsto|'
    r'uparrow|downarrow|Uparrow|Downarrow|updownarrow|'
    r'nearrow|searrow|nwarrow|swarrow|'
    # ── Trig / calculus / analysis ──────────────────────────────────────────
    r'sin|cos|tan|cot|sec|csc|arcsin|arccos|arctan|'
    r'sinh|cosh|tanh|coth|log|ln|exp|'
    # ── Brackets & delimiters ───────────────────────────────────────────────
    r'left|right|lfloor|rfloor|lceil|rceil|langle|rangle|'
    r'bigl|bigr|Bigl|Bigr|biggl|biggr|'
    # ── Environments / structure ────────────────────────────────────────────
    r'begin|end|'
    r'matrix|bmatrix|pmatrix|vmatrix|Vmatrix|Bmatrix|'
    r'cases|array|aligned|align|gather|equation|'
    # ── Text / style ────────────────────────────────────────────────────────
    r'text|mbox|mathrm|mathbf|mathbb|mathcal|mathit|mathsf|mathtt|'
    r'boldsymbol|bm|'
    # ── Over/under decorations ──────────────────────────────────────────────
    r'overline|underline|overbrace|underbrace|widehat|widetilde|'
    r'overrightarrow|xleftarrow|xrightarrow|'
    # ── Combinatorics / algebra ─────────────────────────────────────────────
    r'binom|choose|pmod|mod|gcd|lcm|'
    # ── Chemistry (mhchem-style plain text) ────────────────────────────────
    r'ce|chem|chemfig|'
    # ── Engineering / signals ───────────────────────────────────────────────
    r'laplace|fourier|ztransform|'
    r'degree|celsius|fahrenheit|ohm|'
    # ── Misc ────────────────────────────────────────────────────────────────
    r'quad|qquad|,|;|:|!|'
    r'ldots|cdots|vdots|ddots|'
    r'prime|dagger|ddagger|'
    r'infty|aleph|beth|'
    r'hline|vline|cline'
    r')'
)

# Chemical formula pattern: sequences like H2O, CO2, C6H12O6, Fe2O3,
# NaCl, CaCO3, H2SO4 — must have at least one digit or be multi-element
_CHEM_FORMULA  = re.compile(
    r'\b(?:[A-Z][a-z]?\d*){2,}\b|'      # multi-element: NaCl, H2O, Fe2O3
    r'\b[A-Z][a-z]?\d+\b'               # single element with count: H2, O3
)


def _has_math(text: str) -> bool:
    """Quick check: does this string look like it contains math/LaTeX?"""
    if not text:
        return False
    return bool(
        _LATEX_DISPLAY.search(text) or
        _LATEX_INLINE.search(text) or
        _LATEX_CMDS.search(text)
    )


def _strip_latex_delimiters(expr: str) -> str:
    """Remove surrounding $ or $$ delimiters."""
    expr = expr.strip()
    if expr.startswith("$$") and expr.endswith("$$"):
        return expr[2:-2].strip()
    if expr.startswith("$") and expr.endswith("$"):
        return expr[1:-1].strip()
    return expr


# ── Matplotlib helpers ────────────────────────────────────────────────────────

def _mpl_backend() -> None:
    """Ensure a non-interactive Agg backend is active."""
    import matplotlib
    matplotlib.use("Agg")


def render_math_png(expr: str, fontsize: int = 14, dpi: int = 150) -> bytes | None:
    """
    Render a mathematical expression (LaTeX) to PNG bytes.
    Returns None if rendering fails.
    """
    try:
        _mpl_backend()
        import matplotlib.pyplot as plt

        clean = _strip_latex_delimiters(expr)
        # Wrap in $…$ if not already a display environment
        if not (clean.startswith(r"\begin") or clean.startswith(r"\[")):
            latex = f"${clean}$"
        else:
            latex = clean

        fig = plt.figure(figsize=(6, 1.2))
        fig.patch.set_facecolor("white")
        t = fig.text(
            0.5, 0.5, latex,
            ha="center", va="center",
            fontsize=fontsize,
            color="black",
        )
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=dpi,
                    bbox_inches="tight", pad_inches=0.15,
                    facecolor="white")
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        logger.debug(f"Math render failed for '{expr[:40]}': {e}")
        return None


# ── Excel chart extraction ────────────────────────────────────────────────────

def _resolve_ref(ws, ref) -> list:
    """Resolve an openpyxl Reference object to a flat list of values."""
    if ref is None:
        return []
    vals: list = []
    try:
        for row in ws.iter_rows(
            min_row=ref.min_row, max_row=ref.max_row,
            min_col=ref.min_col, max_col=ref.max_col,
            values_only=True,
        ):
            for v in row:
                if v is not None:
                    vals.append(v)
    except Exception:
        pass
    return vals


def _chart_title(chart) -> str:
    """Best-effort chart title extraction."""
    try:
        t = chart.title
        if t is None:
            return ""
        if isinstance(t, str):
            return t
        # openpyxl Title object
        if hasattr(t, "tx") and t.tx:
            tx = t.tx
            if hasattr(tx, "rich") and tx.rich:
                parts = []
                for para in tx.rich.paragraphs:
                    for run in para.runs:
                        if run.t:
                            parts.append(run.t)
                return "".join(parts)
            if hasattr(tx, "strRef") and tx.strRef and tx.strRef.f:
                return tx.strRef.f
        return ""
    except Exception:
        return ""


def _series_values(ws, series) -> tuple[list, list, str]:
    """
    Return (categories, values, series_label) from an openpyxl series object.
    Works for BarChart, LineChart, AreaChart, PieChart series.
    """
    cats: list = []
    vals: list = []
    label = ""
    try:
        # Series title / legend entry
        if hasattr(series, "title") and series.title:
            st = series.title
            if hasattr(st, "v") and st.v:
                label = str(st.v)
            elif hasattr(st, "strRef") and st.strRef:
                f = getattr(st.strRef, "f", "")
                label = f.split("!")[-1].strip('"') if "!" in f else f
        # Values
        val_ref = getattr(series, "val", None) or getattr(series, "yVal", None)
        if val_ref and hasattr(val_ref, "numRef") and val_ref.numRef:
            nr = val_ref.numRef
            if hasattr(nr, "ref") and nr.ref:
                from openpyxl.utils import range_to_tuple
                try:
                    ws2 = ws.parent[nr.ref.split("!")[0].strip("'")] \
                        if "!" in nr.ref else ws
                    ref_str = nr.ref.split("!")[-1] if "!" in nr.ref else nr.ref
                    from openpyxl import load_workbook as _lw
                    # Use cached values if available
                    if nr.numCache and nr.numCache.pt:
                        vals = [float(p.v) for p in nr.numCache.pt if p.v is not None]
                    else:
                        from openpyxl.utils.cell import range_boundaries
                        mn_col, mn_row, mx_col, mx_row = range_boundaries(ref_str)
                        for row in ws.iter_rows(
                            min_row=mn_row, max_row=mx_row,
                            min_col=mn_col, max_col=mx_col,
                            values_only=True,
                        ):
                            for v in row:
                                if isinstance(v, (int, float)):
                                    vals.append(float(v))
                except Exception:
                    pass
        # Categories
        cat_ref = getattr(series, "cat", None) or getattr(series, "xVal", None)
        if cat_ref:
            if hasattr(cat_ref, "numRef") and cat_ref.numRef and cat_ref.numRef.numCache:
                cats = [str(p.v) for p in cat_ref.numRef.numCache.pt if p.v is not None]
            elif hasattr(cat_ref, "strRef") and cat_ref.strRef and cat_ref.strRef.strCache:
                cats = [str(p.v) for p in cat_ref.strRef.strCache.pt if p.v is not None]
    except Exception as e:
        logger.debug(f"Series extraction error: {e}")
    return cats, vals, label


def _render_chart(chart, ws) -> bytes | None:
    """Render a single openpyxl chart object to PNG bytes using matplotlib."""
    try:
        _mpl_backend()
        import matplotlib.pyplot as plt
        import matplotlib.cm as cm
        import numpy as np

        chart_type = type(chart).__name__.lower()  # barchart, linechart, piechart …
        title = _chart_title(chart)
        series_list = list(chart.series)

        if not series_list:
            return None

        fig, ax = plt.subplots(figsize=(7, 4))
        fig.patch.set_facecolor("white")
        ax.set_facecolor("#f8f9fa")
        if title:
            ax.set_title(title, fontsize=11, fontweight="bold", pad=8)

        colors = cm.tab10.colors

        if "pie" in chart_type:
            cats, vals, _ = _series_values(ws, series_list[0])
            if not vals:
                plt.close(fig)
                return None
            labels = cats if cats else [f"Item {i+1}" for i in range(len(vals))]
            ax.pie(vals, labels=labels, autopct="%1.1f%%",
                   colors=colors[:len(vals)], startangle=90)

        elif "scatter" in chart_type or "bubble" in chart_type:
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                x = list(range(len(vals))) if not cats else cats
                try:
                    x = [float(v) for v in x]
                except (ValueError, TypeError):
                    x = list(range(len(vals)))
                if vals:
                    ax.scatter(x, vals, label=label or f"Series {idx+1}",
                               color=colors[idx % len(colors)], s=60, alpha=0.8)
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        elif "line" in chart_type or "area" in chart_type or "stock" in chart_type:
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                x = list(range(len(vals))) if not cats else cats
                if vals:
                    if "area" in chart_type:
                        ax.fill_between(range(len(vals)), vals,
                                        alpha=0.4, color=colors[idx % len(colors)])
                    ax.plot(vals, label=label or f"Series {idx+1}",
                            color=colors[idx % len(colors)], linewidth=2, marker="o",
                            markersize=4)
                    if cats:
                        step = max(1, len(cats) // 10)
                        ax.set_xticks(range(0, len(cats), step))
                        ax.set_xticklabels(
                            [str(cats[i]) for i in range(0, len(cats), step)],
                            rotation=30, fontsize=7, ha="right"
                        )
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        elif "radar" in chart_type:
            cats, vals, _ = _series_values(ws, series_list[0])
            if vals and cats:
                angles = np.linspace(0, 2 * np.pi, len(vals), endpoint=False).tolist()
                angles += angles[:1]
                vals_closed = vals + vals[:1]
                ax_r = fig.add_subplot(111, polar=True)
                ax.remove()
                ax_r.plot(angles, vals_closed, linewidth=1.5, color=colors[0])
                ax_r.fill(angles, vals_closed, alpha=0.25, color=colors[0])
                ax_r.set_xticks(angles[:-1])
                ax_r.set_xticklabels(cats, fontsize=8)

        else:
            # Default: grouped bar chart
            all_cats: list[str] = []
            all_vals: list[tuple[list, str]] = []
            for idx, series in enumerate(series_list):
                cats, vals, label = _series_values(ws, series)
                if vals:
                    all_vals.append((vals, label or f"Series {idx+1}"))
                    if cats and not all_cats:
                        all_cats = [str(c) for c in cats]

            if not all_vals:
                plt.close(fig)
                return None

            n_groups = max(len(v) for v, _ in all_vals)
            n_series = len(all_vals)
            x = np.arange(n_groups)
            width = 0.8 / n_series

            for idx, (vals, label) in enumerate(all_vals):
                offset = (idx - n_series / 2 + 0.5) * width
                bars = ax.bar(x[:len(vals)] + offset, vals, width,
                              label=label, color=colors[idx % len(colors)], alpha=0.85)

            if all_cats:
                step = max(1, len(all_cats) // 12)
                ax.set_xticks(x[::step])
                ax.set_xticklabels(
                    [all_cats[i] for i in range(0, len(all_cats), step)],
                    rotation=30, fontsize=7, ha="right",
                )
            if n_series > 1:
                ax.legend(fontsize=8)
            ax.grid(axis="y", alpha=0.3)

        plt.tight_layout(pad=0.8)
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=130, bbox_inches="tight",
                    facecolor="white")
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        logger.warning(f"Chart render failed [{type(chart).__name__}]: {e}")
        try:
            plt.close("all")
        except Exception:
            pass
        return None


# ── Public API ────────────────────────────────────────────────────────────────

def extract_excel_charts(path: str, sheet_name: str | None = None) -> list[dict]:
    """
    Extract all charts from an Excel worksheet.

    Returns a list of dicts:
        {
          "title": str,
          "png":   bytes,      # rendered PNG image
        }
    Returns an empty list on any failure.
    """
    results: list[dict] = []
    try:
        import openpyxl
        # Load with data_only=True so cached values are available for reference resolution
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet_names = wb.sheetnames if sheet_name is None else [sheet_name]
        for sn in sheet_names:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            charts = getattr(ws, "_charts", [])
            for chart in charts:
                png = _render_chart(chart, ws)
                if png:
                    results.append({
                        "title": _chart_title(chart) or sn,
                        "png": png,
                    })
        wb.close()
    except Exception as e:
        logger.warning(f"Excel chart extraction failed: {e}")
    return results


def extract_excel_formulas(path: str, sheet_name: str | None = None) -> dict[str, dict]:
    """
    Read formula strings from Excel cells (loads WITHOUT data_only).

    Returns:
        { sheet_name: { "A1": "=SUM(B1:B10)", ... } }
    """
    result: dict[str, dict] = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=False)
        names = wb.sheetnames if sheet_name is None else [sheet_name]
        for sn in names:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            sheet_formulas: dict[str, str] = {}
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        sheet_formulas[cell.coordinate] = v
            if sheet_formulas:
                result[sn] = sheet_formulas
        wb.close()
    except Exception as e:
        logger.debug(f"Formula extraction failed: {e}")
    return result


def detect_math_cells(data: dict[str, Any]) -> dict[str, dict[str, bytes]]:
    """
    Scan a dict of {sheet_name: DataFrame} and detect cells with LaTeX/math content.

    Returns:
        { sheet_name: { "R<row>C<col>": <png_bytes> } }
    """
    results: dict[str, dict[str, bytes]] = {}
    for sheet_name, df in data.items():
        sheet_imgs: dict[str, bytes] = {}
        for r_idx, row in enumerate(df.iter_rows(named=False)):
            for c_idx, val in enumerate(row):
                text = str(val) if val is not None else ""
                if _has_math(text):
                    png = render_math_png(text)
                    if png:
                        sheet_imgs[f"R{r_idx}C{c_idx}"] = png
        if sheet_imgs:
            results[sheet_name] = sheet_imgs
    return results


def extract_docx_equations(path: str) -> list[dict]:
    """
    Extract OMML (Office Math Markup Language) equations from a DOCX file.
    Returns list of { "text": str, "png": bytes | None }.
    """
    results: list[dict] = []
    try:
        from lxml import etree
        from zipfile import ZipFile

        NS_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
        NS_M = "http://schemas.openxmlformats.org/officeDocument/2006/math"

        with ZipFile(path) as zf:
            if "word/document.xml" not in zf.namelist():
                return results
            xml_bytes = zf.read("word/document.xml")

        root = etree.fromstring(xml_bytes)

        for omath in root.iter(f"{{{NS_M}}}oMath"):
            # Best-effort text extraction from OMML
            text_parts = []
            for t_elem in omath.iter(f"{{{NS_M}}}t"):
                if t_elem.text:
                    text_parts.append(t_elem.text)
            # Also grab any plain <w:t> inside runs inside the math
            for t_elem in omath.iter(f"{{{NS_W}}}t"):
                if t_elem.text:
                    text_parts.append(t_elem.text)

            raw_text = " ".join(text_parts).strip()
            if not raw_text:
                continue

            # Try to render as LaTeX/math expression
            png = render_math_png(raw_text) if raw_text else None
            results.append({"text": raw_text, "png": png})

    except Exception as e:
        logger.debug(f"DOCX equation extraction failed: {e}")
    return results


def extract_pptx_equations(path: str) -> list[dict]:
    """
    Extract math equations from a PPTX file.
    Returns list of { "slide": int, "text": str, "png": bytes | None }.
    """
    results: list[dict] = []
    try:
        from lxml import etree
        from zipfile import ZipFile

        NS_A14 = "http://schemas.microsoft.com/office/drawing/2010/main"
        NS_M   = "http://schemas.openxmlformats.org/officeDocument/2006/math"

        with ZipFile(path) as zf:
            slide_files = sorted(
                n for n in zf.namelist()
                if n.startswith("ppt/slides/slide") and n.endswith(".xml")
            )
            for slide_idx, slide_file in enumerate(slide_files, 1):
                xml_bytes = zf.read(slide_file)
                root = etree.fromstring(xml_bytes)
                for omath in root.iter(f"{{{NS_M}}}oMath"):
                    text_parts = [
                        t.text for t in omath.iter(f"{{{NS_M}}}t")
                        if t.text
                    ]
                    raw_text = " ".join(text_parts).strip()
                    if raw_text:
                        png = render_math_png(raw_text)
                        results.append({
                            "slide": slide_idx,
                            "text": raw_text,
                            "png": png,
                        })
    except Exception as e:
        logger.debug(f"PPTX equation extraction failed: {e}")
    return results


def enrich_pdf_with_charts(
    doc,  # fitz.Document (already built, will get new pages appended)
    charts: list[dict],
    title_prefix: str = "المخططات البيانية",
    equations: list[dict] | None = None,
) -> None:
    """
    Append a new page (or pages) to a PyMuPDF document containing
    chart images and optional equation images.
    """
    if not charts and not equations:
        return

    try:
        import pymupdf as fitz
        FONT = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
        PAGE_W, PAGE_H = 595, 842  # A4 portrait
        MARGIN = 30
        TITLE_H = 24
        GAP = 12

        page = doc.new_page(width=PAGE_W, height=PAGE_H)
        try:
            page.insert_font(fontname="dvb", fontfile=FONT)
        except Exception:
            pass

        y = MARGIN

        # Section title
        if charts:
            page.draw_rect(
                fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + TITLE_H),
                color=None, fill=(79/255, 70/255, 229/255), overlay=True,
            )
            page.insert_textbox(
                fitz.Rect(MARGIN + 6, y + 4, PAGE_W - MARGIN - 6, y + TITLE_H - 4),
                title_prefix,
                fontname="dvb", fontsize=11, color=(1, 1, 1),
                align=1,  # center
                overlay=True,
            )
            y += TITLE_H + GAP

        for ch in charts:
            png = ch.get("png")
            if not png:
                continue
            ch_title = ch.get("title", "")

            # If we're near the bottom, start a new page
            if y + 180 > PAGE_H - MARGIN:
                page = doc.new_page(width=PAGE_W, height=PAGE_H)
                try:
                    page.insert_font(fontname="dvb", fontfile=FONT)
                except Exception:
                    pass
                y = MARGIN

            # Chart title label
            if ch_title:
                page.insert_textbox(
                    fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + 14),
                    ch_title,
                    fontname="dvb", fontsize=9, color=(0.2, 0.2, 0.2),
                    align=1, overlay=True,
                )
                y += 16

            # Embed chart image
            available_w = PAGE_W - 2 * MARGIN
            img_stream = fitz.open(stream=png, filetype="png")
            img_w = img_stream[0].rect.width
            img_h = img_stream[0].rect.height
            img_stream.close()

            scale = min(available_w / max(img_w, 1), 200 / max(img_h, 1))
            draw_w = img_w * scale
            draw_h = img_h * scale
            x0 = MARGIN + (available_w - draw_w) / 2

            rect = fitz.Rect(x0, y, x0 + draw_w, y + draw_h)
            page.insert_image(rect, stream=png, overlay=True)
            y += draw_h + GAP

        # Equations section
        if equations:
            if y + TITLE_H + 20 > PAGE_H - MARGIN:
                page = doc.new_page(width=PAGE_W, height=PAGE_H)
                try:
                    page.insert_font(fontname="dvb", fontfile=FONT)
                except Exception:
                    pass
                y = MARGIN

            page.draw_rect(
                fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + TITLE_H),
                color=None, fill=(30/255, 100/255, 60/255), overlay=True,
            )
            page.insert_textbox(
                fitz.Rect(MARGIN + 6, y + 4, PAGE_W - MARGIN - 6, y + TITLE_H - 4),
                "المعادلات والصيغ الرياضية",
                fontname="dvb", fontsize=11, color=(1, 1, 1),
                align=1, overlay=True,
            )
            y += TITLE_H + GAP

            for eq in equations:
                png = eq.get("png")
                text = eq.get("text", "")
                if not png and not text:
                    continue

                if y + 50 > PAGE_H - MARGIN:
                    page = doc.new_page(width=PAGE_W, height=PAGE_H)
                    try:
                        page.insert_font(fontname="dvb", fontfile=FONT)
                    except Exception:
                        pass
                    y = MARGIN

                if png:
                    img_stream = fitz.open(stream=png, filetype="png")
                    img_w = img_stream[0].rect.width
                    img_h = img_stream[0].rect.height
                    img_stream.close()
                    available_w = PAGE_W - 2 * MARGIN
                    scale = min(available_w / max(img_w, 1), 60 / max(img_h, 1))
                    draw_w = img_w * scale
                    draw_h = img_h * scale
                    x0 = MARGIN + (available_w - draw_w) / 2
                    rect = fitz.Rect(x0, y, x0 + draw_w, y + draw_h)
                    page.insert_image(rect, stream=png, overlay=True)
                    y += draw_h + 4
                else:
                    page.insert_textbox(
                        fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + 18),
                        text,
                        fontname="dvb", fontsize=8, color=(0.1, 0.1, 0.5),
                        align=0, overlay=True,
                    )
                    y += 20

    except Exception as e:
        logger.warning(f"enrich_pdf_with_charts failed: {e}")


def charts_to_html_blocks(charts: list[dict], equations: list[dict] | None = None) -> str:
    """
    Convert extracted charts and equations to HTML <img> tags (base64 PNG).
    Returns an HTML string fragment ready to inject into an HTML page.
    """
    import base64
    parts: list[str] = []

    if charts:
        parts.append(
            '<div style="margin:24px 0;"><h3 style="font-family:sans-serif;'
            'color:#4F46E5;border-bottom:2px solid #4F46E5;padding-bottom:6px;">'
            '📊 المخططات البيانية</h3>'
        )
        for ch in charts:
            png = ch.get("png")
            title = ch.get("title", "")
            if not png:
                continue
            b64 = base64.b64encode(png).decode()
            parts.append(
                f'<figure style="margin:16px 0;text-align:center;">'
                f'<img src="data:image/png;base64,{b64}" '
                f'style="max-width:100%;border:1px solid #e2e8f0;border-radius:8px;'
                f'box-shadow:0 2px 8px rgba(0,0,0,.08);" alt="{title}"/>'
                f'{"<figcaption style=\"font-family:sans-serif;font-size:13px;color:#64748b;margin-top:6px;\">" + title + "</figcaption>" if title else ""}'
                f'</figure>'
            )
        parts.append('</div>')

    if equations:
        parts.append(
            '<div style="margin:24px 0;"><h3 style="font-family:sans-serif;'
            'color:#059669;border-bottom:2px solid #059669;padding-bottom:6px;">'
            '∑ المعادلات والصيغ الرياضية</h3>'
        )
        for eq in equations:
            png = eq.get("png")
            text = eq.get("text", "")
            if png:
                b64 = base64.b64encode(png).decode()
                parts.append(
                    f'<figure style="margin:12px 0;text-align:center;">'
                    f'<img src="data:image/png;base64,{b64}" '
                    f'style="max-height:60px;" alt="{text}"/>'
                    f'</figure>'
                )
            elif text:
                parts.append(
                    f'<div style="font-family:monospace;font-size:13px;'
                    f'background:#f1f5f9;padding:8px 12px;border-radius:6px;'
                    f'margin:8px 0;">{text}</div>'
                )
        parts.append('</div>')

    return "\n".join(parts)
